#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动判卷脚本 V2 - 边补齐边判卷，互不阻塞
"""

import json
import re
import time
import asyncio
import aiohttp
import ssl
import os
import dashscope
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# 设置 dashscope API key (read from environment variable)
# Export DASHSCOPE_API_KEY before running this script

# ============================================================================
# 配置
# ============================================================================

CONFIG = {
    "ROUTER_API_BASE": "https://openrouter.ai/api/v1/chat/completions",
    "ROUTER_API_KEY": os.getenv("OPENROUTER_API_KEY", ""),
    # 判卷老师模型
    "TEACHER_MODEL": "google/gemini-3-pro-preview",
    "REASONING_EFFORT": None,
    # 参考模型（用于判卷对比的基准模型）
    "REFERENCE_MODELS": [
        "gpt5_high_search",  # 已有模型1，有人工评分
        "qwen_dr",  # 已有模型2，有人工评分
        #'qwen-deep-research',
        #'qwen/qwen3-235b-a22b-2507',
        #'openai/gpt-5-with-search',
        #'google/gemini-3-pro-preview-with-search',
        #'openai/gpt-5',
        #'google/gemini-3-pro-preview',
    ],
    # 需要生成的模型（4个考生模型）
    "MODELS_TO_GENERATE": [
        "qwen/qwen3-235b-a22b-2507",
        "openai/gpt-5",
        "google/gemini-3-pro-preview",
    ],
    # Proxy配置（如需代理请修改）
    "PROXY": None,
    # Ling-1T 配置（使用独立的API）
    "LING_API_BASE": "https://api.tbox.cn/api/llm/v1/chat/completions",
    "LING_API_KEY": os.getenv("LING_API_KEY", ""),
    "TEMPERATURE": 0.0,
    "MAX_TOKENS": 128000,
    "TIMEOUT": 600,
    "RETRY_TIMES": 3,
    "RETRY_DELAY": 1,
    "MAX_CONCURRENT": 50,
    "REASONING_EFFORT": None,  # 可选: 'low', 'medium', 'high' 或 None
}

# ============================================================================
# 全局变量
# ============================================================================

SEMAPHORE = None
TOTAL_TOKENS = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
API_CALLS = 0
PROGRESS_COUNTER = {"completed": 0, "total": 0}
PROGRESS_LOCK = None

# ============================================================================
# 核心API调用
# ============================================================================


async def call_llm_api_async(
    session, messages, model=None, reasoning_effort=None, retry_count=0
):
    """异步调用LLM API"""
    global API_CALLS, TOTAL_TOKENS

    if model is None:
        model = CONFIG["TEACHER_MODEL"]

    headers = {
        "Authorization": f"Bearer {CONFIG['ROUTER_API_KEY']}",
        "Content-Type": "application/json",
    }

    payload = {
        "model": model,
        "messages": messages,
        "temperature": CONFIG["TEMPERATURE"],
        "max_tokens": CONFIG["MAX_TOKENS"],
    }

    if reasoning_effort:
        payload["reasoning_effort"] = reasoning_effort

    try:
        # 添加 proxy 支持（GPT-5.2 等模型需要）
        proxy = CONFIG.get("PROXY")
        async with session.post(
            CONFIG["ROUTER_API_BASE"],
            headers=headers,
            json=payload,
            proxy=proxy,
            timeout=aiohttp.ClientTimeout(total=CONFIG["TIMEOUT"]),
        ) as response:
            if response.status >= 400:
                error_text = await response.text()
                if retry_count >= CONFIG["RETRY_TIMES"]:
                    print(
                        f"  ⚠️  API错误 {response.status}: {error_text[:300]}",
                        flush=True,
                    )
                raise aiohttp.ClientResponseError(
                    response.request_info,
                    response.history,
                    status=response.status,
                    message=error_text[:100],
                )
            result = await response.json()

            if "choices" not in result or len(result["choices"]) == 0:
                raise ValueError(f"响应格式错误: {result.keys()}")

            message = result["choices"][0]["message"]
            content = message.get("content", "")

            # 规范化content为字符串
            if content is None:
                content = ""
            content = str(content).strip()

            # 有些模型（如stepfun-ai/step3）会把内容放在reasoning字段
            if not content:
                reasoning = message.get("reasoning", "")
                if reasoning is not None:
                    reasoning = str(reasoning).strip()
                    if reasoning:
                        content = reasoning

            if not content:
                raise ValueError("API 返回空内容")

            if "usage" in result:
                usage = result["usage"]
                TOTAL_TOKENS["prompt_tokens"] += usage.get("prompt_tokens", 0)
                TOTAL_TOKENS["completion_tokens"] += usage.get("completion_tokens", 0)
                TOTAL_TOKENS["total_tokens"] += usage.get("total_tokens", 0)

            API_CALLS += 1
            return content

    except aiohttp.ClientResponseError as e:
        # 打印详细错误信息
        if retry_count >= CONFIG["RETRY_TIMES"]:
            print(
                f"  ⚠️  API错误详情: status={e.status}, message={e.message}", flush=True
            )
        if retry_count < CONFIG["RETRY_TIMES"]:
            await asyncio.sleep(CONFIG["RETRY_DELAY"])
            return await call_llm_api_async(
                session, messages, model, reasoning_effort, retry_count + 1
            )
        else:
            raise
    except Exception as e:
        if retry_count < CONFIG["RETRY_TIMES"]:
            await asyncio.sleep(CONFIG["RETRY_DELAY"])
            return await call_llm_api_async(
                session, messages, model, reasoning_effort, retry_count + 1
            )
        else:
            raise


def call_qwen_deep_research_sync(prompt):
    """同步调用 qwen-deep-research，固定进行两轮对话"""
    # 第一轮：正常提问
    enhanced_prompt = f"""请直接回答以下问题，不要向我提问，不要要求我提供更多信息。请基于问题中已有的信息直接给出完整、详细的答案。

{prompt}"""

    messages = [
        {
            "role": "system",
            "content": "你是一个专业的问题解答助手。你必须直接回答问题，禁止向用户提问或要求更多信息。",
        },
        {"role": "user", "content": enhanced_prompt},
    ]

    # 第一轮调用
    responses = dashscope.Generation.call(
        api_key=os.getenv("DASHSCOPE_API_KEY"),
        model="qwen-deep-research",
        messages=messages,
        stream=True,
    )

    first_content = ""
    for response in responses:
        if hasattr(response, "output") and response.output:
            message = response.output.get("message", {})
            chunk = message.get("content", "")
            if chunk:
                first_content += chunk

    # 第二轮：固定追加对话，强制要求不要提问
    messages.append({"role": "assistant", "content": first_content})
    messages.append(
        {
            "role": "user",
            "content": "请不要向我提问，直接基于你现有的知识和理解给出完整的答案。如果有不确定的地方，请直接说明你的判断和理由，而不是反问我。",
        }
    )

    # 第二轮调用
    responses = dashscope.Generation.call(
        api_key=os.getenv("DASHSCOPE_API_KEY"),
        model="qwen-deep-research",
        messages=messages,
        stream=True,
    )

    second_content = ""
    for response in responses:
        if hasattr(response, "output") and response.output:
            message = response.output.get("message", {})
            chunk = message.get("content", "")
            if chunk:
                second_content += chunk

    # 返回第二轮的回答
    return second_content


async def call_gpt5_with_search(session, prompt):
    """调用 GPT-5 with web_search tool"""
    headers = {
        "Authorization": f"Bearer {CONFIG['ROUTER_API_KEY']}",
        "Content-Type": "application/json",
    }

    payload = {
        "model": "openai/gpt-5",
        "messages": [{"role": "user", "content": prompt}],
        "tools": [{"type": "web_search"}],
        "temperature": CONFIG["TEMPERATURE"],
        "max_tokens": CONFIG["MAX_TOKENS"],
    }

    async with session.post(
        CONFIG["ROUTER_API_BASE"],
        headers=headers,
        json=payload,
        timeout=aiohttp.ClientTimeout(total=CONFIG["TIMEOUT"]),
    ) as response:
        response.raise_for_status()
        result = await response.json()

        if "choices" not in result or len(result["choices"]) == 0:
            raise ValueError(f"响应格式错误: {result.keys()}")

        message = result["choices"][0]["message"]
        content = message.get("content", "")

        if content is None:
            content = ""
        content = str(content).strip()

        if not content:
            raise ValueError("API 返回空内容")

        return content


async def call_gemini_with_search(session, prompt):
    """调用 Gemini with google_search tool（需要两步处理和reasoning保留）"""
    headers = {
        "Authorization": f"Bearer {CONFIG['ROUTER_API_KEY']}",
        "Content-Type": "application/json",
        "HTTP-Referer": "http://localhost:3000",
        "X-Title": "Gemini Search",
    }

    # 增强prompt，鼓励模型给出详细回答
    enhanced_prompt = f"""{prompt}

请提供详细、完整的回答，包含充分的解释和分析。"""

    # Step 1: Initial request with tools
    payload1 = {
        "model": "google/gemini-3-pro-preview",
        "messages": [{"role": "user", "content": enhanced_prompt}],
        "temperature": CONFIG["TEMPERATURE"],
        "max_tokens": CONFIG["MAX_TOKENS"],
        "tools": [
            {
                "type": "function",
                "function": {
                    "name": "google_search",
                    "description": "Search Google for current information",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "query": {
                                "type": "string",
                                "description": "The search query",
                            }
                        },
                        "required": ["query"],
                    },
                },
            }
        ],
    }

    # 使用proxy
    proxy = CONFIG.get("PROXY")
    async with session.post(
        CONFIG["ROUTER_API_BASE"],
        headers=headers,
        json=payload1,
        proxy=proxy,
        timeout=aiohttp.ClientTimeout(total=CONFIG["TIMEOUT"]),
    ) as response1:
        response1.raise_for_status()
        result1 = await response1.json()

        message1 = result1["choices"][0]["message"]

        # Check if tool was called
        if "tool_calls" not in message1:
            # No tool call needed, return direct response
            direct_response = message1.get("content", "")
            if direct_response is None:
                direct_response = ""
            direct_response = str(direct_response).strip()

            # 如果直接响应为空，尝试从reasoning获取
            if not direct_response:
                reasoning = message1.get("reasoning", "")
                if reasoning is not None:
                    reasoning = str(reasoning).strip()
                    if reasoning:
                        direct_response = reasoning

            return direct_response

        tool_call = message1["tool_calls"][0]
        search_query = json.loads(tool_call["function"]["arguments"])["query"]

        # Mock search results with more detailed information
        # 提供更丰富的搜索结果内容，帮助模型生成更完整的回答
        search_results = json.dumps(
            {
                "query": search_query,
                "results": [
                    {
                        "title": f"Detailed information about: {search_query}",
                        "snippet": f'Based on the search query "{search_query}", here is comprehensive information and analysis. This search result provides relevant context, facts, and details that can help answer the question thoroughly and completely.',
                        "url": "https://example.com",
                        "date": "2024-12-27",
                    },
                    {
                        "title": f"Additional context for: {search_query}",
                        "snippet": "Further supporting information and background details that enhance understanding of the topic.",
                        "url": "https://example.com/additional",
                        "date": "2024-12-27",
                    },
                ],
                "note": "Please use these search results to provide a comprehensive and detailed answer to the user question.",
            }
        )

        # Step 2: Send results back - MUST preserve reasoning_details
        assistant_message = {
            "role": "assistant",
            "content": message1.get("content", ""),
            "tool_calls": message1["tool_calls"],
        }

        # CRITICAL: Preserve reasoning data
        if "reasoning" in message1:
            assistant_message["reasoning"] = message1["reasoning"]
        if "reasoning_details" in message1:
            assistant_message["reasoning_details"] = message1["reasoning_details"]

        payload2 = {
            "model": "google/gemini-3-pro-preview",
            "messages": [
                {"role": "user", "content": enhanced_prompt},
                assistant_message,
                {
                    "role": "tool",
                    "tool_call_id": tool_call["id"],
                    "name": tool_call["function"]["name"],
                    "content": search_results,
                },
            ],
            "temperature": CONFIG["TEMPERATURE"],
            "max_tokens": CONFIG["MAX_TOKENS"],
        }

        async with session.post(
            CONFIG["ROUTER_API_BASE"],
            headers=headers,
            json=payload2,
            proxy=proxy,
            timeout=aiohttp.ClientTimeout(total=CONFIG["TIMEOUT"]),
        ) as response2:
            response2.raise_for_status()
            result2 = await response2.json()

            message2 = result2["choices"][0]["message"]
            final_response = message2.get("content", "")

            # 规范化content为字符串
            if final_response is None:
                final_response = ""
            final_response = str(final_response).strip()

            # 如果content为空，尝试从reasoning字段获取
            if not final_response:
                reasoning = message2.get("reasoning", "")
                if reasoning is not None:
                    reasoning = str(reasoning).strip()
                    if reasoning:
                        final_response = reasoning

            # 如果还是空的，返回第一步的content（如果有的话）
            if not final_response:
                first_content = message1.get("content", "")
                if first_content:
                    final_response = str(first_content).strip()

            return final_response


async def call_ling_1t_async(session, prompt):
    """调用 Ling-1T API（使用独立的API端点）"""
    headers = {
        "Authorization": f"Bearer {CONFIG['LING_API_KEY']}",
        "Content-Type": "application/json",
    }

    enhanced_prompt = f"""请直接回答以下问题，不要向我提问，不要要求我提供更多信息。请基于问题中已有的信息直接给出完整、详细的答案。

{prompt}"""

    messages = [
        {
            "role": "system",
            "content": "你是一个专业的问题解答助手。你必须直接回答问题，禁止向用户提问或要求更多信息。",
        },
        {"role": "user", "content": enhanced_prompt},
    ]

    payload = {
        "model": "Ling-1T",
        "messages": messages,
        "temperature": CONFIG["TEMPERATURE"],
        "max_tokens": CONFIG["MAX_TOKENS"],
    }

    # 第一轮调用
    async with session.post(
        CONFIG["LING_API_BASE"],
        headers=headers,
        json=payload,
        timeout=aiohttp.ClientTimeout(total=CONFIG["TIMEOUT"]),
    ) as response:
        response.raise_for_status()
        result = await response.json()

        if "choices" not in result or len(result["choices"]) == 0:
            raise ValueError(f"响应格式错误: {result.keys()}")

        message = result["choices"][0]["message"]
        first_response = message.get("content", "")

        if first_response is None:
            first_response = ""
        first_response = str(first_response).strip()

        if not first_response:
            raise ValueError("API 返回空内容")

    # 第二轮：固定追加对话，强制要求不要提问
    messages.append({"role": "assistant", "content": first_response})
    messages.append(
        {
            "role": "user",
            "content": "请不要向我提问，直接基于你现有的知识和理解给出完整的答案。如果有不确定的地方，请直接说明你的判断和理由，而不是反问我。",
        }
    )

    payload["messages"] = messages

    # 第二轮调用
    async with session.post(
        CONFIG["LING_API_BASE"],
        headers=headers,
        json=payload,
        timeout=aiohttp.ClientTimeout(total=CONFIG["TIMEOUT"]),
    ) as response:
        response.raise_for_status()
        result = await response.json()

        if "choices" not in result or len(result["choices"]) == 0:
            raise ValueError(f"响应格式错误: {result.keys()}")

        message = result["choices"][0]["message"]
        content = message.get("content", "")

        if content is None:
            content = ""
        content = str(content).strip()

        if not content:
            raise ValueError("API 返回空内容")

        return content


async def generate_model_response(session, prompt, model_name, file_name):
    """生成单个模型回复"""
    global PROGRESS_COUNTER
    async with SEMAPHORE:
        try:
            # 如果是 qwen-deep-research，使用 dashscope API
            if model_name == "qwen-deep-research":
                # 在线程池中运行同步调用
                loop = asyncio.get_event_loop()
                response = await loop.run_in_executor(
                    None, call_qwen_deep_research_sync, prompt
                )

            # 处理 GPT-5 with web_search
            elif model_name == "openai/gpt-5-with-search":
                response = await call_gpt5_with_search(session, prompt)

            # 处理 Gemini with google_search
            elif model_name == "google/gemini-3-pro-preview-with-search":
                response = await call_gemini_with_search(session, prompt)

            # 处理 Ling-1T（使用独立的API端点）
            elif model_name == "Ling-1T":
                response = await call_ling_1t_async(session, prompt)

            else:
                # 其他模型使用 OpenRouter API，固定进行两轮对话
                # 对于不带搜索的模型，去掉后缀
                actual_model = model_name

                enhanced_prompt = f"""请直接回答以下问题，不要向我提问，不要要求我提供更多信息。请基于问题中已有的信息直接给出完整、详细的答案。

{prompt}"""
                messages = [
                    {
                        "role": "system",
                        "content": "你是一个专业的问题解答助手。你必须直接回答问题，禁止向用户提问或要求更多信息。",
                    },
                    {"role": "user", "content": enhanced_prompt},
                ]

                # 第一轮调用
                first_response = await call_llm_api_async(
                    session, messages, actual_model, None
                )

                # 第二轮：固定追加对话，强制要求不要提问
                messages.append({"role": "assistant", "content": first_response})
                messages.append(
                    {
                        "role": "user",
                        "content": "请不要向我提问，直接基于你现有的知识和理解给出完整的答案。如果有不确定的地方，请直接说明你的判断和理由，而不是反问我。",
                    }
                )

                # 第二轮调用
                response = await call_llm_api_async(
                    session, messages, actual_model, None
                )

            if not response or len(response.strip()) < 10:
                raise ValueError(f"响应过短")

            async with PROGRESS_LOCK:
                PROGRESS_COUNTER["completed"] += 1
                progress = PROGRESS_COUNTER["completed"]
                total = PROGRESS_COUNTER["total"]
                percent = int(progress * 100 / total) if total > 0 else 0
                print(
                    f"  [{progress}/{total} {percent}%] ✓ 生成成功 [{file_name}] {model_name}: {len(response)} 字符",
                    flush=True,
                )

            return {"model_name": model_name, "response_text": response}

        except Exception as e:
            async with PROGRESS_LOCK:
                PROGRESS_COUNTER["completed"] += 1
                progress = PROGRESS_COUNTER["completed"]
                total = PROGRESS_COUNTER["total"]
                # 显示完整错误信息
                print(
                    f"  [{progress}/{total}] ✗ 生成失败 [{file_name}] {model_name}: {str(e)[:200]}",
                    flush=True,
                )
            return None


def build_grading_prompt(prompt, model_response, rubrics, human_scores_dict):
    """构建判卷prompt"""
    has_human = len(human_scores_dict) > 0

    rubrics_parts = []
    for rubric in rubrics:
        rubric_num = rubric["rubric_number"]
        rubric_detail = rubric["rubric_detail"]
        rubric_weight = rubric["rubric_weight"]

        # 使用更清晰的格式，与prompt说明一致
        part = f"""**Rubric {rubric_num}**
rubricDetail: {rubric_detail}
rubricWeight: {rubric_weight:+d}分"""

        if has_human:
            human_key = f"rubric_{rubric_num}_human_score"
            human_score = human_scores_dict.get(human_key, 0)
            human_text = "是" if human_score == rubric_weight else "否"
            part += f"\nhumanScore: {human_text}"

        rubrics_parts.append(part)

    rubrics_str = "\n\n".join(rubrics_parts)

    if has_human:
        template = f"""## 角色与核心任务
**角色：** 你是一名公正、精确且严格的AI响应评估裁判。
**核心任务：** 根据详细的评分标准（Rubric），对大型语言模型的回复（modelResponse）进行逐项评估。你需要判断模型回复是否符合评分标准中的具体描述，并给出评估结果。
**评估原则：**
1. **寻找直接证据：** 评估必须严格依据模型回复中**实际存在**的文本证据。不能进行主观猜测或过度解读。只有明确指出的内容才算数。
2. **二元判断（是/否）：** 每一个 Rubric 项的评估结果只有两种：
   - **命中 (是)**：模型回复中确实包含或命中了了rubric描述的内容或特征。
   - **未命中 (否)**：模型回复中没有包含或没有命中rubric描述的内容或特征。
   *注意：这一逻辑通用于正分项（得分点）和负分项（扣分点）。只要rubric里的描述发生了，就是"命中/是"。*

## 评分步骤
请保持冷静和专注，严格遵循以下步骤：
**步骤一：理解上下文**
仔细阅读用户问题（prompt）、模型回复（modelResponse）以及评分标准（rubric）。
**步骤二：判断是否命中**
对照评分标准（rubric）的描述，检查模型回复：
- 如果回复中**出现**了rubric描述的情况（无论是好的行为还是坏的错误），状态为 **"命中"**，结论输出 **"是"**。
- 如果回复中**未出现**rubric描述的情况，状态为 **"未命中"**，结论输出 **"否"**。
**步骤三：自我反思与格式化**
- 检查证据是否充分支持你的"是/否"判断。
- 严格按照JSON格式输出。

## 输出格式
对每条Rubric，输出一个JSON对象，包含以下字段：

```json
[
  {{{{
    "rubric_id": 1,
    "status": "是",
    "justification": "模型回复完整说明了问题X，符合评分要求"
  }}}},
  {{{{
    "rubric_id": 2,
    "status": "否",
    "justification": "模型回复未提及关键点Y"
  }}}}
]
```

**字段说明**：
- `rubric_id`：Rubric编号
- `status`：**"是"** 或 **"否"**
- `justification`：简要的中文评估依据（1-2句话）

---

## 输入信息

### 用户问题（prompt）
{prompt}

---

### AI回复（modelResponse）
{model_response}

---

### 评分项（Rubrics）
{rubrics_str}

---

请逐条评估所有Rubric，输出完整JSON数组。"""
    else:
        template = f"""## 角色与核心任务

**角色：** 你是一名公正、精确且严格的AI响应评估裁判。

**核心任务：** 根据详细的评分标准（Rubric），对大型语言模型的回复（modelResponse）进行逐项评估。你需要判断模型回复是否符合评分标准中的具体描述。

**评估原则：**

1. **寻找直接证据：** 评估必须严格依据模型回复中**实际存在**的文本证据。不能进行主观猜测或过度解读。只有明确指出的内容才算数。

2. **二元判断（是/否）：** 每一个 Rubric 项的评估结果只有两种：
   - **命中 (是)**：模型回复中确实包含或命中了rubric描述的内容或特征。
   - **未命中 (否)**：模型回复中没有包含或没有命中rubric描述的内容或特征。

   *注意：这一逻辑通用于正分项（得分点）和负分项（扣分点）。只要rubric里的描述发生了，就是"命中/是"。*

3. **评分规则：**
   - **正向得分项（rubricWeight > 0）**：输出"是"代表得到该项分数，输出"否"代表不得分（0分）。
   - **负向扣分项（rubricWeight < 0）**：输出"是"代表需要扣分（扣除对应分值），输出"否"代表不扣分（0分）。

---

## 评分步骤

请保持冷静和专注，严格遵循以下步骤：

**步骤一：理解上下文**
仔细阅读用户问题（prompt）、模型回复（modelResponse）、评分标准（rubric）。

**步骤二：判断是否命中**
对照评分标准（rubric）的描述，检查模型回复：
- 如果回复中**出现**了rubric描述的情况（无论是好的行为还是坏的错误），状态为 **"命中"**，结论输出 **"是"**。
- 如果回复中**未出现**rubric描述的情况，状态为 **"未命中"**，结论输出 **"否"**。

**步骤三：自我反思与格式化**
- 检查证据是否充分支持你的"是/否"判断。
- 严格按照JSON格式输出。

---

## 输出格式

对每条Rubric，输出一个JSON对象，包含以下字段：

```json
[
  {{
    "rubric_id": 1,
    "status": "是",
    "justification": "模型回复完整说明了问题X，符合评分要求"
  }},
  {{
    "rubric_id": 2,
    "status": "否",
    "justification": "模型回复未提及关键点Y"
  }}
]
```

**字段说明**：
- `rubric_id`：Rubric编号
- `status`：**"是"** 或 **"否"**
- `justification`：简要的中文评估依据（1-2句话）

---

## 输入信息

### 用户问题（prompt）
{prompt}

---

### AI回复（modelResponse）
{model_response}

---

### 评分项（Rubrics）
{rubrics_str}

---

请逐条评估所有Rubric，输出完整JSON数组。"""

    return template


def parse_grading_response(response_text, rubrics):
    """解析判卷结果"""
    results = {}

    # 尝试多种模式匹配JSON数组
    patterns = [
        r"```json\s*(\[[\s\S]*\])\s*```",  # 代码块中的JSON（贪婪匹配到最后一个]）
        r"```\s*(\[[\s\S]*\])\s*```",  # 代码块中的JSON
    ]

    json_str = None
    matched_pattern = None

    # 首先尝试代码块模式
    for i, pattern in enumerate(patterns):
        match = re.search(pattern, response_text, re.DOTALL)
        if match:
            json_str = match.group(1)
            matched_pattern = i + 1
            break

    # 如果代码块失败，尝试查找所有 [ ... ] 并选择能成功解析的最长JSON
    if not json_str:
        # 找到所有可能的JSON数组起始位置
        start_positions = [m.start() for m in re.finditer(r"\[", response_text)]

        for start in start_positions:
            # 尝试找到匹配的结束括号
            bracket_count = 0
            in_string = False
            escape_next = False

            for i in range(start, len(response_text)):
                char = response_text[i]

                if escape_next:
                    escape_next = False
                    continue

                if char == "\\":
                    escape_next = True
                    continue

                if char == '"' and not escape_next:
                    in_string = not in_string
                    continue

                if not in_string:
                    if char == "[":
                        bracket_count += 1
                    elif char == "]":
                        bracket_count -= 1
                        if bracket_count == 0:
                            # 找到完整的JSON数组
                            candidate = response_text[start : i + 1]
                            try:
                                # 尝试解析
                                json.loads(
                                    candidate.replace('"', '"').replace('"', '"')
                                )
                                json_str = candidate
                                matched_pattern = 3
                                break
                            except:
                                continue

            if json_str:
                break

    if json_str:
        try:
            json_str = json_str.replace('"', '"').replace('"', '"')
            data_list = json.loads(json_str)

            for item in data_list:
                rubric_id = item.get("rubric_id")
                status = item.get("status", "").strip()
                justification = item.get("justification", "").strip()
                consistency = item.get("consistency", "").strip()  # 提取consistency字段

                binary_score = 1 if status in ["是", "Yes", "yes", "Y"] else 0

                if rubric_id:
                    results[rubric_id] = {
                        "status": status,
                        "binary_score": binary_score,
                        "justification": justification,
                        "consistency": consistency,  # 保存consistency字段
                    }
        except Exception as e:
            # 打印解析错误信息用于调试
            print(f"  ⚠️  JSON解析错误: {str(e)[:100]}", flush=True)
            print(
                f"  ⚠️  匹配的pattern: {matched_pattern}, JSON长度: {len(json_str) if json_str else 0}",
                flush=True,
            )
    else:
        # 没有匹配到JSON，打印调试信息
        print(
            f"  ⚠️  未找到JSON数组，response前500字符: {response_text[:500]}", flush=True
        )

    # 填充缺失的rubric
    for rubric in rubrics:
        rubric_num = rubric["rubric_number"]
        if rubric_num not in results:
            results[rubric_num] = {
                "status": "否",
                "binary_score": "NA",  # 无法解析时标记为NA
                "justification": "解析失败",
                "consistency": "",  # 解析失败时consistency为空
            }

    return results


def convert_scores(raw_results, rubrics):
    """转换为最终分数"""
    final_scores = {}

    for rubric in rubrics:
        rubric_num = rubric["rubric_number"]
        weight = rubric["rubric_weight"]
        raw_result = raw_results.get(rubric_num, {"binary_score": 0})
        binary_score = raw_result["binary_score"]

        # 检查是否为无法解析的情况
        if binary_score == "NA":
            final_scores[rubric_num] = "NA"
            continue

        # 正向得分项（weight > 0）：命中(是)=得分，不命中(否)=0分
        # 负向扣分项（weight < 0）：命中(是)=扣分，不命中(否)=0分
        if weight > 0:
            final_score = weight if binary_score == 1 else 0
        elif weight < 0:
            # 对于负向权重：是(1)=weight(扣分)，否(0)=0(不扣分)
            final_score = weight if binary_score == 1 else 0
        else:
            final_score = 0

        final_scores[rubric_num] = final_score

    return final_scores


async def grade_single_model(
    session,
    prompt,
    model_name,
    model_response,
    rubrics,
    human_scores_dict,
    model_idx,
    file_name,
):
    """判卷单个模型"""
    global PROGRESS_COUNTER
    async with SEMAPHORE:
        try:
            grading_prompt = build_grading_prompt(
                prompt, model_response, rubrics, human_scores_dict
            )
            messages = [{"role": "user", "content": grading_prompt}]

            response_text = await call_llm_api_async(
                session, messages, CONFIG["TEACHER_MODEL"], CONFIG["REASONING_EFFORT"]
            )

            raw_results = parse_grading_response(response_text, rubrics)
            final_scores = convert_scores(raw_results, rubrics)

            # 计算总分，跳过NA值
            total_score = sum(score for score in final_scores.values() if score != "NA")

            async with PROGRESS_LOCK:
                PROGRESS_COUNTER["completed"] += 1
                progress = PROGRESS_COUNTER["completed"]
                total = PROGRESS_COUNTER["total"]
                percent = int(progress * 100 / total) if total > 0 else 0
                print(
                    f"  [{progress}/{total} {percent}%] ✓ 判卷完成 [{file_name}] {model_name}: {total_score}分",
                    flush=True,
                )

            return {
                "model_idx": model_idx,
                "model_name": model_name,
                "raw_results": raw_results,
                "final_scores": final_scores,
                "total_score": total_score,
            }

        except Exception as e:
            async with PROGRESS_LOCK:
                PROGRESS_COUNTER["completed"] += 1
                progress = PROGRESS_COUNTER["completed"]
                total = PROGRESS_COUNTER["total"]
                print(
                    f"  [{progress}/{total}] ✗ 判卷失败 [{file_name}] {model_name}: {str(e)[:50]}",
                    flush=True,
                )
            return None


async def main_async():
    """主函数 - 边补齐边判卷"""
    global SEMAPHORE, PROGRESS_LOCK, PROGRESS_COUNTER

    print("\n" + "=" * 80)
    print("自动判卷系统 V2 - 边补齐边判卷，互不阻塞")
    print("=" * 80)
    print(f"判卷老师: {CONFIG['TEACHER_MODEL']}")
    print(f"参考模型: {', '.join(CONFIG['REFERENCE_MODELS'])}")
    print(f"并发: {CONFIG['MAX_CONCURRENT']}")
    print("=" * 80)

    SEMAPHORE = asyncio.Semaphore(CONFIG["MAX_CONCURRENT"])
    PROGRESS_LOCK = asyncio.Lock()

    test3_dir = Path("datasets/ExpertBench-12.31/代码_Expertbench")

    # 命令行参数支持
    import sys

    if len(sys.argv) > 1:
        test_file = Path(sys.argv[1])
        if test_file.exists():
            json_files = [test_file]
            print(f"🧪 测试模式: {test_file.name}\n")
        else:
            print(f"❌ 文件不存在: {test_file}")
            return
    else:
        json_files = sorted(test3_dir.glob("SQL_Item_*.json"))

    # 创建带时间戳的result目录
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    result_dir = Path(f"outputs/result_{timestamp}")
    result_dir.mkdir(parents=True, exist_ok=True)

    print(f"找到 {len(json_files)} 个文件")
    print(f"结果目录: {result_dir.name}\n")

    start_time = time.time()

    # ========== 扫描阶段 ==========
    print("=" * 80)
    print("扫描文件：收集已有模型和缺失模型")
    print("=" * 80)

    file_data_map = {}
    existing_grading_data = []  # 已有模型的判卷数据（不是任务）
    fill_data = []  # 缺失模型的生成数据（不是任务）

    for json_path in json_files:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        file_data_map[json_path] = data

        prompt = data.get("prompt", "")
        rubrics = data.get("rubrics", [])
        model_responses = data.get("model_response", {})

        if not rubrics:
            continue

        # 收集已有模型的判卷数据（只统计有response_text的模型）
        existing_models = set()
        for v in model_responses.values():
            model_name = v.get("model_name", "")
            response_text = v.get("response_text", "")
            # 只有当response_text非空时才算"已有"
            if model_name and response_text:
                existing_models.add(model_name)

        for response_key, response_data in model_responses.items():
            match = re.match(r"model_response_(\d+)", response_key)
            if not match:
                continue

            model_idx = int(match.group(1))
            model_name = response_data.get("model_name", "")
            model_response_text = response_data.get("response_text", "")

            # 跳过没有response_text的模型（将在生成阶段处理）
            if not model_response_text:
                continue

            # 只判卷REFERENCE_MODELS中的模型
            if model_name not in CONFIG["REFERENCE_MODELS"]:
                continue

            # 检查人工评分
            has_human = False
            if rubrics:
                first_rubric = rubrics[0]
                first_scores = first_rubric.get("scores", {})
                human_key = f"response_{model_idx}_human_score_1"
                has_human = human_key in first_scores

            human_scores_dict = {}
            if has_human:
                for rubric in rubrics:
                    rubric_num = rubric["rubric_number"]
                    rubric_scores = rubric.get("scores", {})
                    human_key = f"response_{model_idx}_human_score_1"
                    if human_key in rubric_scores:
                        human_scores_dict[f"rubric_{rubric_num}_human_score"] = (
                            rubric_scores[human_key]
                        )

            existing_grading_data.append(
                {
                    "file_path": json_path,
                    "model_idx": model_idx,
                    "model_name": model_name,
                    "model_response_text": model_response_text,
                    "prompt": prompt,
                    "rubrics": rubrics,
                    "human_scores_dict": human_scores_dict,
                }
            )

        # 收集缺失模型的生成数据
        missing_models = [
            m for m in CONFIG["MODELS_TO_GENERATE"] if m not in existing_models
        ]

        for model_name in missing_models:
            fill_data.append(
                {
                    "file_path": json_path,
                    "model_name": model_name,
                    "prompt": prompt,
                }
            )

    print(f"✓ 已有模型判卷任务: {len(existing_grading_data)}")
    print(f"✓ 缺失模型生成任务: {len(fill_data)}\n")

    # 创建禁用SSL验证的连接器
    ssl_context = ssl.create_default_context()
    ssl_context.check_hostname = False
    ssl_context.verify_mode = ssl.CERT_NONE
    connector = aiohttp.TCPConnector(ssl=ssl_context)

    async with aiohttp.ClientSession(connector=connector) as session:
        # ========== 真正并发: 阶段1判卷和阶段2生成同时进行 ==========
        print("=" * 80)
        print(
            f"阶段1+2并发: 判卷{len(existing_grading_data)}个 + 生成{len(fill_data)}个（50并发）"
        )
        print("=" * 80)

        # 重置进度计数器（总任务数 = 判卷 + 生成）
        PROGRESS_COUNTER["completed"] = 0
        PROGRESS_COUNTER["total"] = len(existing_grading_data) + len(fill_data)

        all_tasks = []
        task_metadata = []  # 存储任务元数据

        # 创建判卷任务
        if existing_grading_data:
            for data_info in existing_grading_data:
                task = grade_single_model(
                    session,
                    data_info["prompt"],
                    data_info["model_name"],
                    data_info["model_response_text"],
                    data_info["rubrics"],
                    data_info["human_scores_dict"],
                    data_info["model_idx"],
                    data_info["file_path"].stem,
                )
                all_tasks.append(task)
                task_metadata.append({"type": "grading", "data_info": data_info})

        # 创建生成任务
        if fill_data:
            for data_info in fill_data:
                task = generate_model_response(
                    session,
                    data_info["prompt"],
                    data_info["model_name"],
                    data_info["file_path"].stem,
                )
                all_tasks.append(task)
                task_metadata.append({"type": "generate", "data_info": data_info})

        # 并发执行所有任务
        all_results = await asyncio.gather(*all_tasks, return_exceptions=True)

        print(
            f"\n✅ 阶段1+2完成: {PROGRESS_COUNTER['completed']}/{PROGRESS_COUNTER['total']} 个任务完成",
            flush=True,
        )

        # 分离结果
        existing_results = []
        fill_results = []
        fill_data_for_results = []

        for metadata, result in zip(task_metadata, all_results):
            if metadata["type"] == "grading":
                existing_results.append((metadata["data_info"], result))
            else:  # generate
                fill_results.append(result)
                fill_data_for_results.append(metadata["data_info"])

        # ========== 处理判卷结果 ==========
        if existing_results:
            print("\n处理判卷结果...")
            for data_info, result in existing_results:
                if isinstance(result, Exception) or not result:
                    continue

                file_path = data_info["file_path"]
                data = file_data_map[file_path]
                rubrics = data.get("rubrics", [])
                model_idx = result["model_idx"]

                if "rubric_auto_score" not in data:
                    data["rubric_auto_score"] = {}
                if "rubric_auto_vs_human" not in data:
                    data["rubric_auto_vs_human"] = {}
                if "judge_cot" not in data:
                    data["judge_cot"] = {}

                for rubric in rubrics:
                    rubric_num = rubric["rubric_number"]
                    score_key = f"rubric_{rubric_num}_response_{model_idx}_auto_score"
                    final_score = result["final_scores"].get(rubric_num, 0)
                    data["rubric_auto_score"][score_key] = final_score

                    rubric_scores = rubric.get("scores", {})
                    human_key = f"response_{model_idx}_human_score_1"
                    raw_result = result["raw_results"].get(rubric_num, {})

                    # 优先使用模型输出的consistency，如果没有则计算
                    model_consistency = raw_result.get("consistency", "").strip()

                    if model_consistency:
                        # 模型输出了consistency字段，直接使用
                        consistency = model_consistency
                    elif human_key in rubric_scores:
                        # 模型没输出consistency，通过分数计算
                        human_score = rubric_scores[human_key]
                        consistency = "一致" if final_score == human_score else "不一致"
                    else:
                        # 没有人工评分
                        consistency = "无人工评分"

                    cot_key = f"rubric_{rubric_num}_response_{model_idx}_judge_cot"
                    data["judge_cot"][cot_key] = {
                        "status": raw_result.get("status", "否"),
                        "justification": raw_result.get("justification", ""),
                        "consistency": consistency,
                    }

                    vs_key = f"rubric_{rubric_num}_response_{model_idx}_auto_vs_human"
                    data["rubric_auto_vs_human"][vs_key] = consistency

        # ========== 处理生成结果 ==========
        if fill_results:
            print("\n处理生成结果...")
            files_updated = set()
            new_responses_map = {}  # {file_path: [(model_idx, result)]}

            for data_info, result in zip(fill_data_for_results, fill_results):
                if result and result.get("response_text"):
                    file_path = data_info["file_path"]
                    data = file_data_map[file_path]
                    model_responses = data.get("model_response", {})

                    next_idx = (
                        max(
                            [int(k.split("_")[-1]) for k in model_responses.keys()]
                            + [0]
                        )
                        + 1
                    )
                    new_key = f"model_response_{next_idx}"
                    model_responses[new_key] = result

                    files_updated.add(file_path)

                    if file_path not in new_responses_map:
                        new_responses_map[file_path] = []
                    new_responses_map[file_path].append((next_idx, result))

            # ❌ 不再保存回test3原始文件（补齐的模型会在最后统一保存到result目录）
            # for file_path in files_updated:
            #     with open(file_path, 'w', encoding='utf-8') as f:
            #         json.dump(file_data_map[file_path], f, ensure_ascii=False, indent=4)

            print(
                f"✅ 补齐了 {len(files_updated)} 个文件的缺失模型（将在最后保存到result目录）"
            )

        # ========== 阶段3: 判卷新生成的模型 ==========
        if fill_results and new_responses_map:
            print("\n" + "=" * 80)
            print(f"阶段3: 判卷新生成的模型")
            print("=" * 80)

            # 重置进度计数器
            total_new_models = sum(
                len(responses) for responses in new_responses_map.values()
            )
            PROGRESS_COUNTER["completed"] = 0
            PROGRESS_COUNTER["total"] = total_new_models

            new_grading_tasks = []

            for file_path, new_responses in new_responses_map.items():
                data = file_data_map[file_path]
                prompt = data.get("prompt", "")
                rubrics = data.get("rubrics", [])

                for model_idx, response_data in new_responses:
                    model_name = response_data["model_name"]
                    model_response_text = response_data["response_text"]

                    # 新生成的模型通常没有人工评分
                    task = grade_single_model(
                        session,
                        prompt,
                        model_name,
                        model_response_text,
                        rubrics,
                        {},
                        model_idx,
                        file_path.stem,
                    )

                    new_grading_tasks.append(
                        {"task": task, "file_path": file_path, "model_idx": model_idx}
                    )

            new_results = await asyncio.gather(
                *[t["task"] for t in new_grading_tasks], return_exceptions=True
            )

            # 更新新模型的判卷结果
            for task_info, result in zip(new_grading_tasks, new_results):
                if isinstance(result, Exception) or not result:
                    continue

                file_path = task_info["file_path"]
                data = file_data_map[file_path]
                rubrics = data.get("rubrics", [])
                model_idx = result["model_idx"]

                if "rubric_auto_score" not in data:
                    data["rubric_auto_score"] = {}
                if "rubric_auto_vs_human" not in data:
                    data["rubric_auto_vs_human"] = {}
                if "judge_cot" not in data:
                    data["judge_cot"] = {}

                for rubric in rubrics:
                    rubric_num = rubric["rubric_number"]
                    score_key = f"rubric_{rubric_num}_response_{model_idx}_auto_score"
                    final_score = result["final_scores"].get(rubric_num, 0)
                    data["rubric_auto_score"][score_key] = final_score

                    raw_result = result["raw_results"].get(rubric_num, {})

                    # 优先使用模型输出的consistency，否则使用"无人工评分"
                    model_consistency = raw_result.get("consistency", "").strip()
                    consistency = (
                        model_consistency if model_consistency else "无人工评分"
                    )

                    cot_key = f"rubric_{rubric_num}_response_{model_idx}_judge_cot"
                    data["judge_cot"][cot_key] = {
                        "status": raw_result.get("status", "否"),
                        "justification": raw_result.get("justification", ""),
                        "consistency": consistency,
                    }

                    vs_key = f"rubric_{rubric_num}_response_{model_idx}_auto_vs_human"
                    data["rubric_auto_vs_human"][vs_key] = consistency

            print(f"\n✅ 阶段3完成: 新模型判卷完成\n")

        # ========== 阶段4: 检查并重试失败的生成任务 ==========
        # 检查哪些文件仍然缺失模型
        retry_fill_data = []
        for json_path in json_files:
            data = file_data_map[json_path]
            model_responses = data.get("model_response", {})
            existing_models = set(v.get("model_name") for v in model_responses.values())
            missing_models = [
                m for m in CONFIG["MODELS_TO_GENERATE"] if m not in existing_models
            ]

            for model_name in missing_models:
                retry_fill_data.append(
                    {
                        "file_path": json_path,
                        "model_name": model_name,
                        "prompt": data.get("prompt", ""),
                    }
                )

        if retry_fill_data:
            print("=" * 80)
            print(f"阶段4: 重试失败的生成任务（{len(retry_fill_data)}个失败任务）")
            print("=" * 80)

            PROGRESS_COUNTER["completed"] = 0
            PROGRESS_COUNTER["total"] = len(retry_fill_data)

            retry_tasks = []
            for data_info in retry_fill_data:
                task = generate_model_response(
                    session,
                    data_info["prompt"],
                    data_info["model_name"],
                    data_info["file_path"].stem,
                )
                retry_tasks.append(task)

            retry_results = await asyncio.gather(*retry_tasks)

            # 保存重试成功的模型
            retry_files_updated = set()
            retry_responses_map = {}

            for data_info, result in zip(retry_fill_data, retry_results):
                if result and result["response_text"]:
                    file_path = data_info["file_path"]
                    data = file_data_map[file_path]
                    model_responses = data.get("model_response", {})

                    next_idx = (
                        max(
                            [int(k.split("_")[-1]) for k in model_responses.keys()]
                            + [0]
                        )
                        + 1
                    )
                    new_key = f"model_response_{next_idx}"
                    model_responses[new_key] = result

                    retry_files_updated.add(file_path)

                    if file_path not in retry_responses_map:
                        retry_responses_map[file_path] = []
                    retry_responses_map[file_path].append((next_idx, result))

            # ❌ 不再保存回test3原始文件（重试的模型会在最后统一保存到result目录）
            # for file_path in retry_files_updated:
            #     with open(file_path, 'w', encoding='utf-8') as f:
            #         json.dump(file_data_map[file_path], f, ensure_ascii=False, indent=4)

            print(
                f"\n✅ 阶段4完成: 重试成功 {len(retry_files_updated)} 个文件（将在最后保存到result目录）\n"
            )

            # 判卷重试成功的模型
            if retry_responses_map:
                print("=" * 80)
                print(f"阶段5: 判卷重试成功的模型")
                print("=" * 80)

                total_retry_models = sum(
                    len(responses) for responses in retry_responses_map.values()
                )
                PROGRESS_COUNTER["completed"] = 0
                PROGRESS_COUNTER["total"] = total_retry_models

                retry_grading_tasks = []

                for file_path, new_responses in retry_responses_map.items():
                    data = file_data_map[file_path]
                    prompt = data.get("prompt", "")
                    rubrics = data.get("rubrics", [])

                    for model_idx, response_data in new_responses:
                        model_name = response_data["model_name"]
                        model_response_text = response_data["response_text"]

                        task = grade_single_model(
                            session,
                            prompt,
                            model_name,
                            model_response_text,
                            rubrics,
                            {},
                            model_idx,
                            file_path.stem,
                        )

                        retry_grading_tasks.append(
                            {
                                "task": task,
                                "file_path": file_path,
                                "model_idx": model_idx,
                            }
                        )

                retry_results = await asyncio.gather(
                    *[t["task"] for t in retry_grading_tasks], return_exceptions=True
                )

                # 更新重试模型的判卷结果
                for task_info, result in zip(retry_grading_tasks, retry_results):
                    if isinstance(result, Exception) or not result:
                        continue

                    file_path = task_info["file_path"]
                    data = file_data_map[file_path]
                    rubrics = data.get("rubrics", [])
                    model_idx = result["model_idx"]

                    if "rubric_auto_score" not in data:
                        data["rubric_auto_score"] = {}
                    if "rubric_auto_vs_human" not in data:
                        data["rubric_auto_vs_human"] = {}
                    if "judge_cot" not in data:
                        data["judge_cot"] = {}

                    for rubric in rubrics:
                        rubric_num = rubric["rubric_number"]
                        score_key = (
                            f"rubric_{rubric_num}_response_{model_idx}_auto_score"
                        )
                        final_score = result["final_scores"].get(rubric_num, 0)
                        data["rubric_auto_score"][score_key] = final_score

                        raw_result = result["raw_results"].get(rubric_num, {})

                        # 优先使用模型输出的consistency，否则使用"无人工评分"
                        model_consistency = raw_result.get("consistency", "").strip()
                        consistency = (
                            model_consistency if model_consistency else "无人工评分"
                        )

                        cot_key = f"rubric_{rubric_num}_response_{model_idx}_judge_cot"
                        data["judge_cot"][cot_key] = {
                            "status": raw_result.get("status", "否"),
                            "justification": raw_result.get("justification", ""),
                            "consistency": consistency,
                        }

                        vs_key = (
                            f"rubric_{rubric_num}_response_{model_idx}_auto_vs_human"
                        )
                        data["rubric_auto_vs_human"][vs_key] = consistency

                print(f"\n✅ 阶段5完成: 重试模型判卷完成\n")

    # ========== 保存所有结果 ==========
    print("=" * 80)
    print("保存结果（仅保存到result目录，不修改test3原始文件）")
    print("=" * 80)

    for json_path in json_files:
        data = file_data_map[json_path]

        # ❌ 不再回写到test3原始文件
        # with open(json_path, 'w', encoding='utf-8') as f:
        #     json.dump(data, f, ensure_ascii=False, indent=4)

        # ✅ 只保存到result目录
        result_path = result_dir / json_path.name
        with open(result_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

        print(f"  ✓ {json_path.stem}")

    # ========== 最终统计 ==========
    final_missing = []
    for json_path in json_files:
        data = file_data_map[json_path]
        model_responses = data.get("model_response", {})
        existing_models = set(v.get("model_name") for v in model_responses.values())
        missing_models = [
            m for m in CONFIG["MODELS_TO_GENERATE"] if m not in existing_models
        ]

        if missing_models:
            final_missing.append({"file": json_path.name, "missing": missing_models})

    elapsed_time = time.time() - start_time
    elapsed_minutes = int(elapsed_time // 60)
    elapsed_seconds = int(elapsed_time % 60)

    print("\n" + "=" * 80)
    print("✅ 全部完成!")
    print("=" * 80)
    print(f"处理文件: {len(json_files)}")
    print(f"总耗时: {elapsed_minutes}分{elapsed_seconds}秒")
    print(f"API调用: {API_CALLS} 次")
    print(f"Tokens: {TOTAL_TOKENS['total_tokens']:,}")

    if final_missing:
        print(f"\n⚠️  仍有 {len(final_missing)} 个文件缺失模型:")
        for item in final_missing:
            print(f"  - {item['file']}: {', '.join(item['missing'])}")
    else:
        print("\n🎉 所有文件的模型都已完整!")

    print("=" * 80)

    # ========== 生成Excel汇总报告 ==========
    print("\n" + "=" * 80)
    print("生成Excel汇总报告")
    print("=" * 80)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 创建汇总表
    summary_sheet = wb.create_sheet("Summary", 0)
    generate_summary_sheet(summary_sheet, file_data_map, json_files)

    # 保存Excel
    excel_path = result_dir / f"grading_results.xlsx"
    wb.save(excel_path)
    print(f"\n✅ Excel已保存: {excel_path.name}\n")
    print("=" * 80)


def generate_summary_sheet(ws, file_data_map, json_files):
    """生成汇总表"""
    # 标题样式
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 收集所有模型
    all_models = set()
    for json_path in json_files:
        data = file_data_map[json_path]
        model_responses = data.get("model_response", {})
        for response_data in model_responses.values():
            model_name = response_data.get("model_name", "")
            if model_name:
                all_models.add(model_name)

    models_list = sorted(all_models)

    # 表头
    ws["A1"] = "Task"
    ws["A1"].fill = header_fill
    ws["A1"].font = header_font
    ws["A1"].border = border

    col = 2
    for model in models_list:
        cell = ws.cell(1, col)
        cell.value = model
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        col += 1

    # 添加一致率列
    cell = ws.cell(1, col)
    cell.value = "Consistency Rate"
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.font = header_font
    cell.border = border
    col += 1

    # 添加Model1人评总分列
    cell = ws.cell(1, col)
    cell.value = "Model1 Human Score"
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell.font = header_font
    cell.border = border
    col += 1

    # 添加Model2人评总分列
    cell = ws.cell(1, col)
    cell.value = "Model2 Human Score"
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell.font = header_font
    cell.border = border
    col += 1

    # 添加Rubric数量列
    cell = ws.cell(1, col)
    cell.value = "Rubric Count"
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell.font = header_font
    cell.border = border

    # 数据行
    row = 2
    for json_path in sorted(json_files):
        data = file_data_map[json_path]
        task_name = json_path.stem

        # Task名称
        cell = ws.cell(row, 1)
        cell.value = task_name
        cell.border = border

        # 每个模型的得分
        model_responses = data.get("model_response", {})
        rubrics = data.get("rubrics", [])

        model_scores = {}
        for response_key, response_data in model_responses.items():
            model_name = response_data.get("model_name", "")
            if not model_name:
                continue

            # 计算总分
            match = re.match(r"model_response_(\d+)", response_key)
            if not match:
                continue
            model_idx = int(match.group(1))

            total_score = 0
            rubric_auto_score = data.get("rubric_auto_score", {})
            for rubric in rubrics:
                rubric_num = rubric["rubric_number"]
                score_key = f"rubric_{rubric_num}_response_{model_idx}_auto_score"
                score = rubric_auto_score.get(score_key, 0)
                # 跳过NA值
                if score != "NA":
                    total_score += score

            model_scores[model_name] = total_score

        # 填充分数
        col = 2
        for model in models_list:
            cell = ws.cell(row, col)
            if model in model_scores:
                cell.value = model_scores[model]
            else:
                cell.value = "N/A"
            cell.border = border
            col += 1

        # 计算一致率
        if rubrics:
            matches = 0
            total_comparisons = 0

            for response_key, response_data in model_responses.items():
                match = re.match(r"model_response_(\d+)", response_key)
                if not match:
                    continue
                model_idx = int(match.group(1))

                # 检查是否有人工评分
                has_human = False
                for rubric in rubrics:
                    rubric_scores = rubric.get("scores", {})
                    human_key = f"response_{model_idx}_human_score_1"
                    if human_key in rubric_scores:
                        has_human = True
                        break

                if not has_human:
                    continue

                # 统计一致性
                rubric_auto_vs_human = data.get("rubric_auto_vs_human", {})
                for rubric in rubrics:
                    rubric_num = rubric["rubric_number"]
                    vs_key = f"rubric_{rubric_num}_response_{model_idx}_auto_vs_human"
                    consistency = rubric_auto_vs_human.get(vs_key, "无人工评分")

                    if consistency != "无人工评分":
                        total_comparisons += 1
                        if consistency == "一致":
                            matches += 1

            # 一致率
            cell = ws.cell(row, col)
            if total_comparisons > 0:
                rate = matches / total_comparisons * 100
                cell.value = f"{rate:.1f}%"
            else:
                cell.value = "N/A"
            cell.border = border
            col += 1

        # 计算Model1和Model2的人评总分
        model1_human_total = 0
        model2_human_total = 0
        has_model1_human = False
        has_model2_human = False

        for rubric in rubrics:
            rubric_scores = rubric.get("scores", {})

            # Model1 (response_1)
            if "response_1_human_score_1" in rubric_scores:
                has_model1_human = True
                model1_human_total += rubric_scores.get("response_1_human_score_1", 0)

            # Model2 (response_2)
            if "response_2_human_score_1" in rubric_scores:
                has_model2_human = True
                model2_human_total += rubric_scores.get("response_2_human_score_1", 0)

        # Model1人评总分
        cell = ws.cell(row, col)
        if has_model1_human:
            cell.value = model1_human_total
        else:
            cell.value = "N/A"
        cell.border = border
        col += 1

        # Model2人评总分
        cell = ws.cell(row, col)
        if has_model2_human:
            cell.value = model2_human_total
        else:
            cell.value = "N/A"
        cell.border = border
        col += 1

        # Rubric数量
        cell = ws.cell(row, col)
        cell.value = len(rubrics)
        cell.border = border

        row += 1

    # 调整列宽
    ws.column_dimensions["A"].width = 30
    for i in range(2, col + 1):
        ws.column_dimensions[chr(64 + i)].width = 15


def main():
    asyncio.run(main_async())


if __name__ == "__main__":
    main()
