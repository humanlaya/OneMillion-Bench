#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel reporting for omb."""

import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..utils.colors import ExcelColors as GruvboxColors


def _sanitize_sheet_name(name: str) -> str:
    """Sanitize a string for use as an Excel sheet name.

    Excel sheet names must be <= 31 chars and cannot contain []:*?/\\.
    """
    # Remove common prefixes to shorten
    short = name.replace("google/", "").replace("anthropic/", "").replace("openai/", "")
    # Remove invalid chars
    for ch in "[]:*?/\\":
        short = short.replace(ch, "_")
    return short[:31]


def _get_active_judge_models(file_data_map: Dict[Path, Dict[str, Any]], json_files: List[Path]) -> List[str]:
    """Get list of judge models that have scores in the data.

    Args:
        file_data_map: Dictionary mapping file paths to data
        json_files: List of JSON file paths

    Returns:
        List of judge model names with data
    """
    active = set()
    for json_path in json_files:
        data = file_data_map[json_path]
        rubric_auto_score = data.get("rubric_auto_score", {})
        for key in rubric_auto_score:
            if key != "_legacy":
                active.add(key)
    return sorted(active) if active else ["_legacy"]


def _get_judge_scores(data: Dict[str, Any], section: str, judge_model_name: str) -> Dict[str, Any]:
    """Get scores from nested data structure for a specific judge model.

    Args:
        data: File data dictionary
        section: Section name (e.g. 'rubric_auto_score')
        judge_model_name: Judge model name key

    Returns:
        Score dictionary for the specified judge model
    """
    section_data = data.get(section, {})
    return section_data.get(judge_model_name, {})


def _parse_judge_run_name(name: str) -> Tuple[str, Optional[int]]:
    """Parse a judge model name that may contain ::run_N suffix.

    Returns:
        Tuple of (base_name, run_index_or_None)
    """
    if "::run_" in name:
        base, suffix = name.rsplit("::run_", 1)
        try:
            return base, int(suffix)
        except ValueError:
            return name, None
    return name, None


def generate_excel_report(
    file_data_map: Dict[Path, Dict[str, Any]],
    json_files: List[Path],
    output_path: Path,
    token_tracker: Any = None,
) -> None:
    """Generate Excel report with grading results.

    Creates one Performance sheet per judge model.

    Args:
        file_data_map: Dictionary mapping file paths to data
        json_files: List of JSON file paths
        output_path: Path to save Excel file
        token_tracker: Optional TokenTracker instance for cost reporting
    """
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create one summary sheet per judge model
    active_judges = _get_active_judge_models(file_data_map, json_files)
    sheet_idx = 0
    for judge_model_name in active_judges:
        sheet_name = _sanitize_sheet_name(judge_model_name)
        # Avoid duplicate sheet names
        if sheet_name in wb.sheetnames:
            sheet_name = sheet_name[:28] + f"_{sheet_idx}"
        summary_sheet = wb.create_sheet(sheet_name, sheet_idx)
        generate_summary_sheet(summary_sheet, file_data_map, json_files, judge_model_name)
        sheet_idx += 1

    # Create aggregate sheets for repeated judge runs
    run_groups: Dict[str, List[str]] = defaultdict(list)
    for judge_name in active_judges:
        base, run_idx = _parse_judge_run_name(judge_name)
        if run_idx is not None:
            run_groups[base].append(judge_name)

    for base_judge, run_names in run_groups.items():
        if len(run_names) < 2:
            continue
        sheet_name = _sanitize_sheet_name(f"AGG_{base_judge}")
        if sheet_name in wb.sheetnames:
            sheet_name = sheet_name[:28] + f"_{sheet_idx}"
        agg_sheet = wb.create_sheet(sheet_name, sheet_idx)
        _generate_aggregate_sheet(agg_sheet, file_data_map, json_files, run_names, base_judge)
        sheet_idx += 1

    # Create cost breakdown sheet if token_tracker is provided
    if token_tracker:
        cost_sheet = wb.create_sheet("Cost Breakdown", sheet_idx)
        generate_cost_sheet(cost_sheet, token_tracker)

    # Save Excel
    wb.save(output_path)


def generate_summary_sheet(
    ws: Any, file_data_map: Dict[Path, Dict[str, Any]], json_files: List[Path],
    judge_model_name: str = "_legacy",
) -> None:
    """Generate summary sheet with all model scores for a specific judge model.

    Args:
        ws: Worksheet object
        file_data_map: Dictionary mapping file paths to data
        json_files: List of JSON file paths
        judge_model_name: Judge model name to read scores from
    """
    # Style definitions with Gruvbox colors
    header_fill = PatternFill(
        start_color=GruvboxColors.BLUE_NEUTRAL,
        end_color=GruvboxColors.BLUE_NEUTRAL,
        fill_type="solid",
    )
    header_font = Font(bold=True, color=GruvboxColors.FG0, size=11)
    border = Border(
        left=Side(style="thin", color=GruvboxColors.BG3),
        right=Side(style="thin", color=GruvboxColors.BG3),
        top=Side(style="thin", color=GruvboxColors.BG3),
        bottom=Side(style="thin", color=GruvboxColors.BG3),
    )

    # Collect all models
    all_models = set()
    for json_path in json_files:
        data = file_data_map[json_path]
        model_responses = data.get("model_response", {})
        for response_data in model_responses.values():
            model_name = response_data.get("model_name", "")
            if model_name:
                all_models.add(model_name)

    models_list = sorted(all_models)

    # Header row
    ws["A1"] = "Task"
    ws["A1"].fill = header_fill
    ws["A1"].font = header_font
    ws["A1"].border = border

    # Add Lower column
    ws["B1"] = "Lower"
    ws["B1"].fill = PatternFill(
        start_color=GruvboxColors.RED_NEUTRAL,
        end_color=GruvboxColors.RED_NEUTRAL,
        fill_type="solid",
    )
    ws["B1"].font = header_font
    ws["B1"].border = border

    # Add upper column
    ws["C1"] = "Upper"
    ws["C1"].fill = PatternFill(
        start_color=GruvboxColors.GREEN_NEUTRAL,
        end_color=GruvboxColors.GREEN_NEUTRAL,
        fill_type="solid",
    )
    ws["C1"].font = header_font
    ws["C1"].border = border

    col = 4  # Start model columns from column 4 (D)
    for model in models_list:
        cell = ws.cell(1, col)
        cell.value = model
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        col += 1

    # Add consistency rate column with gruvbox green
    cell = ws.cell(1, col)
    cell.value = "Consistency Rate"
    cell.fill = PatternFill(
        start_color=GruvboxColors.GREEN_NEUTRAL,
        end_color=GruvboxColors.GREEN_NEUTRAL,
        fill_type="solid",
    )
    cell.font = header_font
    cell.border = border
    col += 1

    # Add Model1 human score column with gruvbox orange
    cell = ws.cell(1, col)
    cell.value = "Model1 Human Score"
    cell.fill = PatternFill(
        start_color=GruvboxColors.ORANGE_NEUTRAL,
        end_color=GruvboxColors.ORANGE_NEUTRAL,
        fill_type="solid",
    )
    cell.font = header_font
    cell.border = border
    col += 1

    # Add Model2 human score column with gruvbox orange
    cell = ws.cell(1, col)
    cell.value = "Model2 Human Score"
    cell.fill = PatternFill(
        start_color=GruvboxColors.ORANGE_NEUTRAL,
        end_color=GruvboxColors.ORANGE_NEUTRAL,
        fill_type="solid",
    )
    cell.font = header_font
    cell.border = border
    col += 1

    # Add Rubric count column with gruvbox yellow
    cell = ws.cell(1, col)
    cell.value = "Rubric Count"
    cell.fill = PatternFill(
        start_color=GruvboxColors.YELLOW_NEUTRAL,
        end_color=GruvboxColors.YELLOW_NEUTRAL,
        fill_type="solid",
    )
    cell.font = header_font
    cell.border = border

    # Data rows
    last_row = _generate_data_rows(ws, file_data_map, json_files, models_list, border, judge_model_name)

    # Add macro average row
    last_row = _add_macro_average_row(
        ws, file_data_map, json_files, models_list, border, last_row + 1, judge_model_name
    )

    # Add micro average row
    last_row = _add_micro_average_row(
        ws, file_data_map, json_files, models_list, border, last_row + 1, judge_model_name
    )

    # Add per-model consistency rate summary
    _add_model_consistency_summary(ws, file_data_map, json_files, models_list, border, last_row + 2, judge_model_name)

    # Adjust column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12  # lower bound
    ws.column_dimensions["C"].width = 12  # upper bound
    for i in range(4, col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15


def _generate_data_rows(
    ws: Any,
    file_data_map: Dict[Path, Dict[str, Any]],
    json_files: List[Path],
    models_list: List[str],
    border: Border,
    judge_model_name: str = "_legacy",
) -> int:
    """Generate data rows for summary sheet.

    Args:
        ws: Worksheet object
        file_data_map: Dictionary mapping file paths to data
        json_files: List of JSON file paths
        models_list: List of model names
        border: Border style
        judge_model_name: Judge model name to read scores from

    Returns:
        Last row number used
    """
    row = 2
    for json_path in sorted(json_files):
        data = file_data_map[json_path]
        task_name = json_path.stem

        # Task name with gruvbox foreground color
        cell = ws.cell(row, 1)
        cell.value = task_name
        cell.border = border
        cell.font = Font(color=GruvboxColors.FG1)

        # Calculate min and max scores from rubric weights
        rubrics = data.get("rubrics", [])
        task_min_score = 0
        task_max_score = 0
        for rubric in rubrics:
            rubric_weight = rubric.get("rubric_weight", 0)
            if rubric_weight > 0:
                task_max_score += rubric_weight
            elif rubric_weight < 0:
                task_min_score += rubric_weight

        # Add lower bound
        cell = ws.cell(row, 2)
        cell.value = task_min_score
        cell.border = border
        cell.font = Font(color=GruvboxColors.RED if task_min_score < 0 else GruvboxColors.FG2)

        # Add upper bound
        cell = ws.cell(row, 3)
        cell.value = task_max_score
        cell.border = border
        cell.font = Font(color=GruvboxColors.GREEN)

        # Each model's score
        model_responses = data.get("model_response", {})
        rubrics = data.get("rubrics", [])

        model_scores = {}
        for response_key, response_data in model_responses.items():
            model_name = response_data.get("model_name", "")
            if not model_name:
                continue

            # Calculate total score
            match = re.match(r"model_response_(\d+)", response_key)
            if not match:
                continue
            model_idx = int(match.group(1))

            total_score = 0
            rubric_auto_score = _get_judge_scores(data, "rubric_auto_score", judge_model_name)
            for rubric in rubrics:
                rubric_num = rubric["rubric_number"]
                score_key = f"rubric_{rubric_num}_response_{model_idx}_auto_score"
                score = rubric_auto_score.get(score_key, 0)
                # Skip NA values
                if score != "NA":
                    total_score += score

            model_scores[model_name] = total_score

        # Fill scores with gruvbox colors
        col = 4  # Start from column 4 (D) since columns B and C are Min/Max scores
        for model in models_list:
            cell = ws.cell(row, col)
            if model in model_scores:
                cell.value = model_scores[model]
                cell.font = Font(color=GruvboxColors.FG2)
            else:
                cell.value = "N/A"
                cell.font = Font(color=GruvboxColors.GRAY)
            cell.border = border
            col += 1

        # Calculate consistency rate
        _add_consistency_rate(ws, row, col, data, model_responses, rubrics, border, judge_model_name)
        col += 1

        # Add human scores
        _add_human_scores(ws, row, col, rubrics, border)
        col += 2

        # Rubric count
        cell = ws.cell(row, col)
        cell.value = len(rubrics)
        cell.border = border
        cell.font = Font(color=GruvboxColors.FG2)

        row += 1

    return row - 1


def _add_consistency_rate(
    ws: Any,
    row: int,
    col: int,
    data: Dict[str, Any],
    model_responses: Dict[str, Any],
    rubrics: List[Dict[str, Any]],
    border: Border,
    judge_model_name: str = "_legacy",
) -> None:
    """Add consistency rate to worksheet.

    Args:
        ws: Worksheet object
        row: Current row
        col: Current column
        data: Data dictionary
        model_responses: Model responses dictionary
        rubrics: List of rubrics
        border: Border style
        judge_model_name: Judge model name to read scores from
    """
    if rubrics:
        matches = 0
        total_comparisons = 0

        for response_key, response_data in model_responses.items():
            match = re.match(r"model_response_(\d+)", response_key)
            if not match:
                continue
            model_idx = int(match.group(1))

            # Check if has human scores
            has_human = False
            for rubric in rubrics:
                rubric_scores = rubric.get("scores", {})
                human_key = f"response_{model_idx}_human_score_1"
                if human_key in rubric_scores:
                    has_human = True
                    break

            if not has_human:
                continue

            # Count consistency
            rubric_auto_vs_human = _get_judge_scores(data, "rubric_auto_vs_human", judge_model_name)
            for rubric in rubrics:
                rubric_num = rubric["rubric_number"]
                vs_key = f"rubric_{rubric_num}_response_{model_idx}_auto_vs_human"
                consistency = rubric_auto_vs_human.get(vs_key, "无人工评分")

                # Handle both English and Chinese values
                if consistency not in ("无人工评分", "No human score"):
                    total_comparisons += 1
                    if consistency in ("一致", "Consistent"):
                        matches += 1

        # Consistency rate with gruvbox colors
        cell = ws.cell(row, col)
        if total_comparisons > 0:
            rate = matches / total_comparisons * 100
            cell.value = f"{rate:.1f}%"
            # Color based on rate: high = green, medium = yellow, low = red
            if rate >= 80:
                cell.font = Font(color=GruvboxColors.GREEN, bold=True)
            elif rate >= 60:
                cell.font = Font(color=GruvboxColors.YELLOW)
            else:
                cell.font = Font(color=GruvboxColors.RED)
        else:
            cell.value = "N/A"
            cell.font = Font(color=GruvboxColors.GRAY)
        cell.border = border


def _add_human_scores(
    ws: Any, row: int, col: int, rubrics: List[Dict[str, Any]], border: Border
) -> None:
    """Add human scores for Model1 and Model2.

    Args:
        ws: Worksheet object
        row: Current row
        col: Starting column
        rubrics: List of rubrics
        border: Border style
    """
    # Calculate Model1 and Model2 human scores
    model1_human_total = 0
    model2_human_total = 0
    has_model1_human = False
    has_model2_human = False

    for rubric in rubrics:
        rubric_scores = rubric.get("scores", {})

        # Model1 (response_1)
        if "response_1_human_score_1" in rubric_scores:
            has_model1_human = True
            model1_human_total += rubric_scores.get("response_1_human_score_1", 0)

        # Model2 (response_2)
        if "response_2_human_score_1" in rubric_scores:
            has_model2_human = True
            model2_human_total += rubric_scores.get("response_2_human_score_1", 0)

    # Model1 human score with gruvbox colors
    cell = ws.cell(row, col)
    if has_model1_human:
        cell.value = model1_human_total
        cell.font = Font(color=GruvboxColors.FG2)
    else:
        cell.value = "N/A"
        cell.font = Font(color=GruvboxColors.GRAY)
    cell.border = border

    # Model2 human score with gruvbox colors
    cell = ws.cell(row, col + 1)
    if has_model2_human:
        cell.value = model2_human_total
        cell.font = Font(color=GruvboxColors.FG2)
    else:
        cell.value = "N/A"
        cell.font = Font(color=GruvboxColors.GRAY)
    cell.border = border


def _add_macro_average_row(
    ws: Any,
    file_data_map: Dict[Path, Dict[str, Any]],
    json_files: List[Path],
    models_list: List[str],
    border: Border,
    row: int,
    judge_model_name: str = "_legacy",
) -> int:
    """Add macro average row showing average scores for each model.

    Args:
        ws: Worksheet object
        file_data_map: Dictionary mapping file paths to data
        json_files: List of JSON file paths
        models_list: List of model names
        border: Border style
        row: Row number to add the average
        judge_model_name: Judge model name to read scores from

    Returns:
        Last row number used
    """
    # Calculate scores for each model across all tasks
    model_scores = {model: [] for model in models_list}
    model_totals = {model: {"score": 0, "max_score": 0, "count": 0} for model in models_list}

    for json_path in sorted(json_files):
        data = file_data_map[json_path]
        model_responses = data.get("model_response", {})
        rubrics = data.get("rubrics", [])
        rubric_count = len(rubrics)

        # Calculate max score from rubric weights
        # For positive weights: max score is the weight value
        # For negative weights: max score is 0 (best is avoiding the penalty)
        task_max_score = 0
        for rubric in rubrics:
            rubric_weight = rubric.get("rubric_weight", 0)
            if rubric_weight > 0:
                task_max_score += rubric_weight

        for response_key, response_data in model_responses.items():
            model_name = response_data.get("model_name", "")
            if not model_name or model_name not in models_list:
                continue

            # Calculate total score for this model on this task
            match = re.match(r"model_response_(\d+)", response_key)
            if not match:
                continue
            model_idx = int(match.group(1))

            total_score = 0
            rubric_auto_score = _get_judge_scores(data, "rubric_auto_score", judge_model_name)
            valid_rubrics = 0
            for rubric in rubrics:
                rubric_num = rubric["rubric_number"]
                score_key = f"rubric_{rubric_num}_response_{model_idx}_auto_score"
                score = rubric_auto_score.get(score_key, 0)
                # Skip NA values
                if score != "NA":
                    total_score += score
                    valid_rubrics += 1

            if valid_rubrics > 0:
                model_scores[model_name].append(total_score)
                model_totals[model_name]["score"] += total_score
                model_totals[model_name]["max_score"] += task_max_score
                model_totals[model_name]["count"] += 1

    # Add macro average row
    cell = ws.cell(row, 1)
    cell.value = "Macro Average (%)"
    cell.fill = PatternFill(
        start_color=GruvboxColors.PURPLE_NEUTRAL,
        end_color=GruvboxColors.PURPLE_NEUTRAL,
        fill_type="solid",
    )
    cell.font = Font(bold=True, color=GruvboxColors.FG0, size=11)
    cell.border = border

    # Leave lower/upper columns empty for average row
    for empty_col in [2, 3]:
        cell = ws.cell(row, empty_col)
        cell.value = ""
        cell.border = border

    col = 4  # Start from column 4 (D) for model scores
    for model in models_list:
        cell = ws.cell(row, col)
        stats = model_totals[model]

        if stats["count"] > 0 and stats["max_score"] > 0:
            # Calculate percentage: (total_score / total_max_score) * 100
            percentage = (stats["score"] / stats["max_score"]) * 100
            cell.value = f"{percentage:.1f}%"
            cell.font = Font(color=GruvboxColors.AQUA, bold=True)
        else:
            cell.value = "N/A"
            cell.font = Font(color=GruvboxColors.GRAY)

        cell.border = border
        col += 1

    return row


def _add_micro_average_row(
    ws: Any,
    file_data_map: Dict[Path, Dict[str, Any]],
    json_files: List[Path],
    models_list: List[str],
    border: Border,
    row: int,
    judge_model_name: str = "_legacy",
) -> int:
    """Add micro average row showing average task accuracy for each model.

    Args:
        ws: Worksheet object
        file_data_map: Dictionary mapping file paths to data
        json_files: List of JSON file paths
        models_list: List of model names
        border: Border style
        row: Row number to add the average
        judge_model_name: Judge model name to read scores from

    Returns:
        Last row number used
    """
    # Calculate task accuracies for each model
    model_task_accuracies = {model: [] for model in models_list}

    for json_path in sorted(json_files):
        data = file_data_map[json_path]
        model_responses = data.get("model_response", {})
        rubrics = data.get("rubrics", [])

        # Calculate max score from rubric weights
        task_max_score = 0
        for rubric in rubrics:
            rubric_weight = rubric.get("rubric_weight", 0)
            if rubric_weight > 0:
                task_max_score += rubric_weight

        for response_key, response_data in model_responses.items():
            model_name = response_data.get("model_name", "")
            if not model_name or model_name not in models_list:
                continue

            # Calculate total score for this model on this task
            match = re.match(r"model_response_(\d+)", response_key)
            if not match:
                continue
            model_idx = int(match.group(1))

            total_score = 0
            rubric_auto_score = _get_judge_scores(data, "rubric_auto_score", judge_model_name)
            valid_rubrics = 0
            for rubric in rubrics:
                rubric_num = rubric["rubric_number"]
                score_key = f"rubric_{rubric_num}_response_{model_idx}_auto_score"
                score = rubric_auto_score.get(score_key, 0)
                # Skip NA values
                if score != "NA":
                    total_score += score
                    valid_rubrics += 1

            if valid_rubrics > 0 and task_max_score > 0:
                task_accuracy = total_score / task_max_score
                model_task_accuracies[model_name].append(task_accuracy)

    # Add micro average row
    cell = ws.cell(row, 1)
    cell.value = "Micro Average (%)"
    cell.fill = PatternFill(
        start_color=GruvboxColors.PURPLE_NEUTRAL,
        end_color=GruvboxColors.PURPLE_NEUTRAL,
        fill_type="solid",
    )
    cell.font = Font(bold=True, color=GruvboxColors.FG0, size=11)
    cell.border = border

    # Leave upper/lower columns empty for average row
    for empty_col in [2, 3]:
        cell = ws.cell(row, empty_col)
        cell.value = ""
        cell.border = border

    col = 4  # Start from column 4 (D) for model scores
    for model in models_list:
        cell = ws.cell(row, col)
        task_accuracies = model_task_accuracies[model]

        if len(task_accuracies) > 0:
            # Calculate average task accuracy and convert to percentage
            avg_accuracy = sum(task_accuracies) / len(task_accuracies)
            percentage = avg_accuracy * 100
            cell.value = f"{percentage:.1f}%"
            cell.font = Font(color=GruvboxColors.AQUA, bold=True)
        else:
            cell.value = "N/A"
            cell.font = Font(color=GruvboxColors.GRAY)

        cell.border = border
        col += 1

    return row


def _add_model_consistency_summary(
    ws: Any,
    file_data_map: Dict[Path, Dict[str, Any]],
    json_files: List[Path],
    models_list: List[str],
    border: Border,
    start_row: int,
    judge_model_name: str = "_legacy",
) -> None:
    """Add per-model consistency rate summary.

    Args:
        ws: Worksheet object
        file_data_map: Dictionary mapping file paths to data
        json_files: List of JSON file paths
        models_list: List of model names
        border: Border style
        start_row: Starting row for the summary
        judge_model_name: Judge model name to read scores from
    """
    # Calculate per-model consistency rates
    model_consistency_stats = {}

    for model_name in models_list:
        model_matches = 0
        model_total = 0

        for json_path in json_files:
            data = file_data_map[json_path]
            model_responses = data.get("model_response", {})
            rubrics = data.get("rubrics", [])
            rubric_auto_vs_human = _get_judge_scores(data, "rubric_auto_vs_human", judge_model_name)

            # Find this model's response index
            model_idx = None
            for response_key, response_data in model_responses.items():
                if response_data.get("model_name", "") == model_name:
                    match = re.match(r"model_response_(\d+)", response_key)
                    if match:
                        model_idx = int(match.group(1))
                        break

            if model_idx is None:
                continue

            # Count consistency for this model in this file
            for rubric in rubrics:
                rubric_num = rubric["rubric_number"]
                vs_key = f"rubric_{rubric_num}_response_{model_idx}_auto_vs_human"
                consistency = rubric_auto_vs_human.get(vs_key, "无人工评分")

                # Handle both English and Chinese values
                if consistency not in ("无人工评分", "No human score"):
                    model_total += 1
                    if consistency in ("一致", "Consistent"):
                        model_matches += 1

        model_consistency_stats[model_name] = {
            "matches": model_matches,
            "total": model_total,
            "rate": (model_matches / model_total * 100) if model_total > 0 else None,
        }

    # Add summary header with gruvbox purple
    row = start_row
    cell = ws.cell(row, 1)
    cell.value = "Per-Model Consistency Rate"
    cell.fill = PatternFill(
        start_color=GruvboxColors.PURPLE_NEUTRAL,
        end_color=GruvboxColors.PURPLE_NEUTRAL,
        fill_type="solid",
    )
    cell.font = Font(bold=True, color=GruvboxColors.FG0, size=11)
    cell.border = border

    # Leave lower/upper columns empty for consistency summary row
    for empty_col in [2, 3]:
        cell = ws.cell(row, empty_col)
        cell.value = ""
        cell.border = border

    # Add model consistency rates
    col = 4  # Start from column 4 (D) for model scores
    for model in models_list:
        cell = ws.cell(row, col)
        stats = model_consistency_stats[model]

        if stats["rate"] is not None:
            cell.value = f"{stats['rate']:.1f}%"
            # Color based on rate
            if stats["rate"] >= 80:
                cell.font = Font(color=GruvboxColors.GREEN, bold=True)
            elif stats["rate"] >= 60:
                cell.font = Font(color=GruvboxColors.YELLOW)
            else:
                cell.font = Font(color=GruvboxColors.RED)
        else:
            cell.value = "N/A"
            cell.font = Font(color=GruvboxColors.GRAY)

        cell.border = border
        col += 1

    # Add overall consistency rate
    row += 1
    cell = ws.cell(row, 1)
    cell.value = "Overall Consistency Rate"
    cell.fill = PatternFill(
        start_color=GruvboxColors.AQUA_NEUTRAL,
        end_color=GruvboxColors.AQUA_NEUTRAL,
        fill_type="solid",
    )
    cell.font = Font(bold=True, color=GruvboxColors.FG0, size=11)
    cell.border = border

    # Calculate overall consistency
    total_matches = sum(stats["matches"] for stats in model_consistency_stats.values())
    total_comparisons = sum(stats["total"] for stats in model_consistency_stats.values())

    cell = ws.cell(row, 2)
    if total_comparisons > 0:
        overall_rate = total_matches / total_comparisons * 100
        cell.value = f"{overall_rate:.1f}%"
        if overall_rate >= 80:
            cell.font = Font(color=GruvboxColors.GREEN, bold=True, size=12)
        elif overall_rate >= 60:
            cell.font = Font(color=GruvboxColors.YELLOW, size=12)
        else:
            cell.font = Font(color=GruvboxColors.RED, size=12)
    else:
        cell.value = "N/A"
        cell.font = Font(color=GruvboxColors.GRAY)
    cell.border = border


def generate_cost_sheet(ws: Any, token_tracker: Any) -> None:
    """Generate cost breakdown sheet with detailed model costs.

    Args:
        ws: Worksheet object
        token_tracker: TokenTracker instance
    """
    # Style definitions with Gruvbox colors
    header_fill = PatternFill(
        start_color=GruvboxColors.BLUE_NEUTRAL,
        end_color=GruvboxColors.BLUE_NEUTRAL,
        fill_type="solid",
    )
    header_font = Font(bold=True, color=GruvboxColors.FG0, size=11)
    border = Border(
        left=Side(style="thin", color=GruvboxColors.BG3),
        right=Side(style="thin", color=GruvboxColors.BG3),
        top=Side(style="thin", color=GruvboxColors.BG3),
        bottom=Side(style="thin", color=GruvboxColors.BG3),
    )

    # Get cost summary from token tracker
    summary = token_tracker.get_summary()
    cost_breakdown = summary["cost_breakdown"]
    total_cost = summary["total_cost"]

    # Header row
    headers = [
        "Model Name",
        "Prompt Tokens",
        "Completion Tokens",
        "Total Tokens",
        "Prompt Cost ($)",
        "Completion Cost ($)",
        "Total Cost ($)",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    row = 2
    for model_name, costs in sorted(cost_breakdown.items()):
        # Model name
        cell = ws.cell(row, 1)
        cell.value = model_name
        cell.border = border
        cell.font = Font(color=GruvboxColors.FG1)

        # Prompt tokens
        cell = ws.cell(row, 2)
        cell.value = costs["prompt_tokens"]
        cell.border = border
        cell.font = Font(color=GruvboxColors.FG2)
        cell.alignment = Alignment(horizontal="right")

        # Completion tokens
        cell = ws.cell(row, 3)
        cell.value = costs["completion_tokens"]
        cell.border = border
        cell.font = Font(color=GruvboxColors.FG2)
        cell.alignment = Alignment(horizontal="right")

        # Total tokens
        cell = ws.cell(row, 4)
        cell.value = costs["prompt_tokens"] + costs["completion_tokens"]
        cell.border = border
        cell.font = Font(color=GruvboxColors.FG2)
        cell.alignment = Alignment(horizontal="right")

        # Prompt cost
        cell = ws.cell(row, 5)
        cell.value = f"${costs['prompt_cost']:.4f}"
        cell.border = border
        cell.font = Font(color=GruvboxColors.AQUA)
        cell.alignment = Alignment(horizontal="right")

        # Completion cost
        cell = ws.cell(row, 6)
        cell.value = f"${costs['completion_cost']:.4f}"
        cell.border = border
        cell.font = Font(color=GruvboxColors.AQUA)
        cell.alignment = Alignment(horizontal="right")

        # Total cost
        cell = ws.cell(row, 7)
        cell.value = f"${costs['total_cost']:.4f}"
        cell.border = border
        cell.font = Font(color=GruvboxColors.GREEN, bold=True)
        cell.alignment = Alignment(horizontal="right")

        row += 1

    # Add total row
    row += 1
    cell = ws.cell(row, 1)
    cell.value = "TOTAL"
    cell.fill = PatternFill(
        start_color=GruvboxColors.ORANGE_NEUTRAL,
        end_color=GruvboxColors.ORANGE_NEUTRAL,
        fill_type="solid",
    )
    cell.font = Font(bold=True, color=GruvboxColors.FG0, size=12)
    cell.border = border

    # Total tokens
    total_tokens_data = summary["total_tokens"]

    cell = ws.cell(row, 2)
    cell.value = total_tokens_data["prompt"]
    cell.fill = PatternFill(
        start_color=GruvboxColors.BG2,
        end_color=GruvboxColors.BG2,
        fill_type="solid",
    )
    cell.font = Font(bold=True, color=GruvboxColors.FG0)
    cell.border = border
    cell.alignment = Alignment(horizontal="right")

    cell = ws.cell(row, 3)
    cell.value = total_tokens_data["completion"]
    cell.fill = PatternFill(
        start_color=GruvboxColors.BG2,
        end_color=GruvboxColors.BG2,
        fill_type="solid",
    )
    cell.font = Font(bold=True, color=GruvboxColors.FG0)
    cell.border = border
    cell.alignment = Alignment(horizontal="right")

    cell = ws.cell(row, 4)
    cell.value = total_tokens_data["total"]
    cell.fill = PatternFill(
        start_color=GruvboxColors.BG2,
        end_color=GruvboxColors.BG2,
        fill_type="solid",
    )
    cell.font = Font(bold=True, color=GruvboxColors.FG0)
    cell.border = border
    cell.alignment = Alignment(horizontal="right")

    # Total cost
    cell = ws.cell(row, 7)
    cell.value = f"${total_cost:.4f}"
    cell.fill = PatternFill(
        start_color=GruvboxColors.ORANGE_NEUTRAL,
        end_color=GruvboxColors.ORANGE_NEUTRAL,
        fill_type="solid",
    )
    cell.font = Font(bold=True, color=GruvboxColors.FG0, size=12)
    cell.border = border
    cell.alignment = Alignment(horizontal="right")

    # Adjust column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 18


def _generate_aggregate_sheet(
    ws: Any,
    file_data_map: Dict[Path, Dict[str, Any]],
    json_files: List[Path],
    run_names: List[str],
    base_judge: str,
) -> None:
    """Generate an aggregate sheet averaging scores across repeated judge runs.

    For each task and model, computes mean and std of scores across runs.

    Args:
        ws: Worksheet object
        file_data_map: Dictionary mapping file paths to data
        json_files: List of JSON file paths
        run_names: List of run-indexed judge names (e.g. ["judge::run_1", "judge::run_2"])
        base_judge: Base judge model name
    """
    header_fill = PatternFill(
        start_color=GruvboxColors.BLUE_NEUTRAL,
        end_color=GruvboxColors.BLUE_NEUTRAL,
        fill_type="solid",
    )
    header_font = Font(bold=True, color=GruvboxColors.FG0, size=11)
    border = Border(
        left=Side(style="thin", color=GruvboxColors.BG3),
        right=Side(style="thin", color=GruvboxColors.BG3),
        top=Side(style="thin", color=GruvboxColors.BG3),
        bottom=Side(style="thin", color=GruvboxColors.BG3),
    )

    # Collect all models
    all_models = set()
    for json_path in json_files:
        data = file_data_map[json_path]
        for response_data in data.get("model_response", {}).values():
            model_name = response_data.get("model_name", "")
            if model_name:
                all_models.add(model_name)
    models_list = sorted(all_models)

    # Headers: Task | model1_mean | model1_std | model2_mean | model2_std | ...
    ws["A1"] = "Task"
    ws["A1"].fill = header_fill
    ws["A1"].font = header_font
    ws["A1"].border = border

    col = 2
    for model in models_list:
        cell = ws.cell(1, col)
        cell.value = f"{model} (mean)"
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        col += 1
        cell = ws.cell(1, col)
        cell.value = f"{model} (std)"
        cell.fill = PatternFill(
            start_color=GruvboxColors.YELLOW_NEUTRAL,
            end_color=GruvboxColors.YELLOW_NEUTRAL,
            fill_type="solid",
        )
        cell.font = header_font
        cell.border = border
        col += 1

    # Data rows
    row = 2
    for json_path in sorted(json_files):
        data = file_data_map[json_path]
        task_name = json_path.stem
        model_responses = data.get("model_response", {})
        rubrics = data.get("rubrics", [])

        cell = ws.cell(row, 1)
        cell.value = task_name
        cell.border = border
        cell.font = Font(color=GruvboxColors.FG1)

        col = 2
        for model_name in models_list:
            # Find model_idx for this model
            model_idx = None
            for rk, rv in model_responses.items():
                if rv.get("model_name", "") == model_name:
                    match = re.match(r"model_response_(\d+)", rk)
                    if match:
                        model_idx = int(match.group(1))
                        break

            scores = []
            if model_idx is not None:
                for run_name in run_names:
                    run_scores = _get_judge_scores(data, "rubric_auto_score", run_name)
                    total = 0
                    valid = False
                    for rubric in rubrics:
                        rn = rubric["rubric_number"]
                        sk = f"rubric_{rn}_response_{model_idx}_auto_score"
                        s = run_scores.get(sk, 0)
                        if s != "NA":
                            total += s
                            valid = True
                    if valid:
                        scores.append(total)

            # Mean
            cell = ws.cell(row, col)
            if scores:
                mean_val = sum(scores) / len(scores)
                cell.value = round(mean_val, 2)
            else:
                cell.value = "N/A"
            cell.border = border
            col += 1

            # Std
            cell = ws.cell(row, col)
            if len(scores) > 1:
                mean_val = sum(scores) / len(scores)
                variance = sum((v - mean_val) ** 2 for v in scores) / (len(scores) - 1)
                cell.value = round(variance ** 0.5, 2)
            else:
                cell.value = "N/A"
            cell.border = border
            col += 1

        row += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 30
    max_col = 2 + len(models_list) * 2
    for i in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18


def _build_repeated_judge_aggregate(
    run_data_list: List[Dict[str, Any]], models_list: List[str]
) -> Dict[str, Any]:
    """Build aggregate performance data across repeated judge runs.

    Args:
        run_data_list: List of performance data dicts, one per run
        models_list: List of model names

    Returns:
        Aggregated performance dict with mean/std for key metrics
    """
    # Aggregate macro_averages, micro_averages, per_model_consistency_rate
    result = {"models": models_list}

    for metric_key in ["macro_averages", "micro_averages", "per_model_consistency_rate"]:
        mean_vals = {}
        std_vals = {}
        for model_name in models_list:
            values = []
            for rd in run_data_list:
                val_str = rd.get(metric_key, {}).get(model_name, "N/A")
                if val_str != "N/A":
                    try:
                        values.append(float(val_str.rstrip("%")))
                    except (ValueError, AttributeError):
                        pass
            if values:
                mean_val = sum(values) / len(values)
                mean_vals[model_name] = f"{mean_val:.1f}%"
                if len(values) > 1:
                    variance = sum((v - mean_val) ** 2 for v in values) / (len(values) - 1)
                    std_vals[model_name] = f"±{variance ** 0.5:.1f}%"
                else:
                    std_vals[model_name] = "N/A"
            else:
                mean_vals[model_name] = "N/A"
                std_vals[model_name] = "N/A"
        result[metric_key] = mean_vals
        result[f"{metric_key}_std"] = std_vals

    # Aggregate task_scores: mean and std per task per model
    if run_data_list and run_data_list[0].get("task_scores"):
        num_tasks = len(run_data_list[0]["task_scores"])
        agg_task_scores = []
        for t_idx in range(num_tasks):
            task_entry = {
                "task_name": run_data_list[0]["task_scores"][t_idx]["task_name"],
                "rubric_count": run_data_list[0]["task_scores"][t_idx]["rubric_count"],
                "model_scores_mean": {},
                "model_scores_std": {},
            }
            for model_name in models_list:
                scores = []
                for rd in run_data_list:
                    if t_idx < len(rd["task_scores"]):
                        s = rd["task_scores"][t_idx].get("model_scores", {}).get(model_name)
                        if s is not None:
                            scores.append(s)
                if scores:
                    mean_val = sum(scores) / len(scores)
                    task_entry["model_scores_mean"][model_name] = round(mean_val, 2)
                    if len(scores) > 1:
                        variance = sum((v - mean_val) ** 2 for v in scores) / (len(scores) - 1)
                        task_entry["model_scores_std"][model_name] = round(variance ** 0.5, 2)
                    else:
                        task_entry["model_scores_std"][model_name] = None
                else:
                    task_entry["model_scores_mean"][model_name] = None
                    task_entry["model_scores_std"][model_name] = None
            agg_task_scores.append(task_entry)
        result["task_scores"] = agg_task_scores

    return result


def generate_json_report(
    file_data_map: Dict[Path, Dict[str, Any]],
    json_files: List[Path],
    output_path: Path,
    token_tracker: Any = None,
) -> None:
    """Generate JSON report with grading results.

    Args:
        file_data_map: Dictionary mapping file paths to data
        json_files: List of JSON file paths
        output_path: Path to save JSON file
        token_tracker: Optional TokenTracker instance for cost reporting
    """
    # Collect all models
    all_models = set()
    for json_path in json_files:
        data = file_data_map[json_path]
        model_responses = data.get("model_response", {})
        for response_data in model_responses.values():
            model_name = response_data.get("model_name", "")
            if model_name:
                all_models.add(model_name)

    models_list = sorted(all_models)

    # Build performance data per judge model
    active_judges = _get_active_judge_models(file_data_map, json_files)
    performance_data = {}
    for judge_model_name in active_judges:
        performance_data[judge_model_name] = _build_performance_data(
            file_data_map, json_files, models_list, judge_model_name
        )

    # Group repeated judge runs and build structured output
    run_groups: Dict[str, List[str]] = defaultdict(list)
    for judge_name in active_judges:
        base, run_idx = _parse_judge_run_name(judge_name)
        if run_idx is not None:
            run_groups[base].append(judge_name)

    # If there are repeated runs, restructure performance data
    if run_groups:
        grouped_performance = {}
        used_run_names = set()
        for base_judge, run_names in run_groups.items():
            if len(run_names) < 2:
                continue
            runs_data = {}
            for rn in sorted(run_names):
                _, idx = _parse_judge_run_name(rn)
                runs_data[f"run_{idx}"] = performance_data[rn]
                used_run_names.add(rn)

            # Compute aggregate
            aggregate = _build_repeated_judge_aggregate(
                [performance_data[rn] for rn in sorted(run_names)], models_list
            )
            grouped_performance[base_judge] = {
                "runs": runs_data,
                "aggregate": aggregate,
            }

        # Keep non-run judges as-is
        for judge_name in active_judges:
            if judge_name not in used_run_names:
                grouped_performance[judge_name] = performance_data[judge_name]

        performance_data = grouped_performance

    # Build cost breakdown data
    cost_data = {}
    if token_tracker:
        summary = token_tracker.get_summary()
        cost_data = {
            "cost_breakdown": summary["cost_breakdown"],
            "total_tokens": summary["total_tokens"],
            "total_cost": summary["total_cost"],
            "model_usage": summary["model_usage"],
        }

    # Combine into final report
    report = {
        "performance": performance_data,
        "cost_breakdown": cost_data,
    }

    # Save JSON file
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=4)


def _build_performance_data(
    file_data_map: Dict[Path, Dict[str, Any]],
    json_files: List[Path],
    models_list: List[str],
    judge_model_name: str = "_legacy",
) -> Dict[str, Any]:
    """Build performance data for JSON report.

    Args:
        file_data_map: Dictionary mapping file paths to data
        json_files: List of JSON file paths
        models_list: List of model names
        judge_model_name: Judge model name to read scores from

    Returns:
        Dictionary containing performance metrics
    """
    # Task-level scores
    task_scores = []

    for json_path in sorted(json_files):
        data = file_data_map[json_path]
        task_name = json_path.stem
        model_responses = data.get("model_response", {})
        rubrics = data.get("rubrics", [])

        task_data = {
            "task_name": task_name,
            "rubric_count": len(rubrics),
            "model_scores": {},
            "consistency_rate": None,
            "model1_human_score": None,
            "model2_human_score": None,
        }

        # Calculate each model's score
        for response_key, response_data in model_responses.items():
            model_name = response_data.get("model_name", "")
            if not model_name:
                continue

            match = re.match(r"model_response_(\d+)", response_key)
            if not match:
                continue
            model_idx = int(match.group(1))

            total_score = 0
            rubric_auto_score = _get_judge_scores(data, "rubric_auto_score", judge_model_name)
            for rubric in rubrics:
                rubric_num = rubric["rubric_number"]
                score_key = f"rubric_{rubric_num}_response_{model_idx}_auto_score"
                score = rubric_auto_score.get(score_key, 0)
                if score != "NA":
                    total_score += score

            task_data["model_scores"][model_name] = total_score

        # Calculate consistency rate
        if rubrics:
            matches = 0
            total_comparisons = 0

            for response_key, response_data in model_responses.items():
                match = re.match(r"model_response_(\d+)", response_key)
                if not match:
                    continue
                model_idx = int(match.group(1))

                # Check if has human scores
                has_human = False
                for rubric in rubrics:
                    rubric_scores = rubric.get("scores", {})
                    human_key = f"response_{model_idx}_human_score_1"
                    if human_key in rubric_scores:
                        has_human = True
                        break

                if not has_human:
                    continue

                # Count consistency
                rubric_auto_vs_human = _get_judge_scores(data, "rubric_auto_vs_human", judge_model_name)
                for rubric in rubrics:
                    rubric_num = rubric["rubric_number"]
                    vs_key = f"rubric_{rubric_num}_response_{model_idx}_auto_vs_human"
                    consistency = rubric_auto_vs_human.get(vs_key, "无人工评分")

                    if consistency not in ("无人工评分", "No human score"):
                        total_comparisons += 1
                        if consistency in ("一致", "Consistent"):
                            matches += 1

            if total_comparisons > 0:
                task_data["consistency_rate"] = f"{matches / total_comparisons * 100:.1f}%"

        # Human scores for Model1 and Model2
        model1_human_total = 0
        model2_human_total = 0
        has_model1_human = False
        has_model2_human = False

        for rubric in rubrics:
            rubric_scores = rubric.get("scores", {})

            if "response_1_human_score_1" in rubric_scores:
                has_model1_human = True
                model1_human_total += rubric_scores.get("response_1_human_score_1", 0)

            if "response_2_human_score_1" in rubric_scores:
                has_model2_human = True
                model2_human_total += rubric_scores.get("response_2_human_score_1", 0)

        if has_model1_human:
            task_data["model1_human_score"] = model1_human_total
        if has_model2_human:
            task_data["model2_human_score"] = model2_human_total

        task_scores.append(task_data)

    # Calculate macro averages
    model_totals = {
        model: {
            "score": 0,
            "max_score": 0,
            "min_score": 0,
            "count": 0,
            "task_accuracies": [],
        }
        for model in models_list
    }

    for json_path in sorted(json_files):
        data = file_data_map[json_path]
        model_responses = data.get("model_response", {})
        rubrics = data.get("rubrics", [])

        # Calculate max and min scores from rubric weights
        task_max_score = 0
        task_min_score = 0
        for rubric in rubrics:
            rubric_weight = rubric.get("rubric_weight", 0)
            if rubric_weight > 0:
                task_max_score += rubric_weight
            elif rubric_weight < 0:
                task_min_score += rubric_weight

        for response_key, response_data in model_responses.items():
            model_name = response_data.get("model_name", "")
            if not model_name or model_name not in models_list:
                continue

            match = re.match(r"model_response_(\d+)", response_key)
            if not match:
                continue
            model_idx = int(match.group(1))

            total_score = 0
            rubric_auto_score = _get_judge_scores(data, "rubric_auto_score", judge_model_name)
            valid_rubrics = 0
            for rubric in rubrics:
                rubric_num = rubric["rubric_number"]
                score_key = f"rubric_{rubric_num}_response_{model_idx}_auto_score"
                score = rubric_auto_score.get(score_key, 0)
                if score != "NA":
                    total_score += score
                    valid_rubrics += 1

            if valid_rubrics > 0:
                model_totals[model_name]["score"] += total_score
                model_totals[model_name]["max_score"] += task_max_score
                model_totals[model_name]["min_score"] += task_min_score
                model_totals[model_name]["count"] += 1

                # Calculate task accuracy for Micro Average
                if task_max_score > 0:
                    task_accuracy = total_score / task_max_score
                    model_totals[model_name]["task_accuracies"].append(task_accuracy)

    macro_averages = {}
    for model_name in models_list:
        stats = model_totals[model_name]
        if stats["count"] > 0 and stats["max_score"] > 0:
            percentage = (stats["score"] / stats["max_score"]) * 100
            macro_averages[model_name] = f"{percentage:.1f}%"
        else:
            macro_averages[model_name] = "N/A"

    # Calculate micro averages
    micro_averages = {}
    for model_name in models_list:
        stats = model_totals[model_name]
        task_accuracies = stats.get("task_accuracies", [])
        if len(task_accuracies) > 0:
            avg_accuracy = sum(task_accuracies) / len(task_accuracies)
            percentage = avg_accuracy * 100
            micro_averages[model_name] = f"{percentage:.1f}%"
        else:
            micro_averages[model_name] = "N/A"

    # Calculate per-model consistency rates
    model_consistency_stats = {}

    for model_name in models_list:
        model_matches = 0
        model_total = 0

        for json_path in json_files:
            data = file_data_map[json_path]
            model_responses = data.get("model_response", {})
            rubrics = data.get("rubrics", [])
            rubric_auto_vs_human = _get_judge_scores(data, "rubric_auto_vs_human", judge_model_name)

            # Find this model's response index
            model_idx = None
            for response_key, response_data in model_responses.items():
                if response_data.get("model_name", "") == model_name:
                    match = re.match(r"model_response_(\d+)", response_key)
                    if match:
                        model_idx = int(match.group(1))
                        break

            if model_idx is None:
                continue

            # Count consistency for this model in this file
            for rubric in rubrics:
                rubric_num = rubric["rubric_number"]
                vs_key = f"rubric_{rubric_num}_response_{model_idx}_auto_vs_human"
                consistency = rubric_auto_vs_human.get(vs_key, "无人工评分")

                if consistency not in ("无人工评分", "No human score"):
                    model_total += 1
                    if consistency in ("一致", "Consistent"):
                        model_matches += 1

        if model_total > 0:
            rate = model_matches / model_total * 100
            model_consistency_stats[model_name] = f"{rate:.1f}%"
        else:
            model_consistency_stats[model_name] = "N/A"

    # Calculate overall consistency rate
    overall_consistency_rate = None
    # Recalculate properly
    all_matches = 0
    all_total = 0
    for model_name in models_list:
        for json_path in json_files:
            data = file_data_map[json_path]
            model_responses = data.get("model_response", {})
            rubrics = data.get("rubrics", [])
            rubric_auto_vs_human = _get_judge_scores(data, "rubric_auto_vs_human", judge_model_name)

            model_idx = None
            for response_key, response_data in model_responses.items():
                if response_data.get("model_name", "") == model_name:
                    match = re.match(r"model_response_(\d+)", response_key)
                    if match:
                        model_idx = int(match.group(1))
                        break

            if model_idx is None:
                continue

            for rubric in rubrics:
                rubric_num = rubric["rubric_number"]
                vs_key = f"rubric_{rubric_num}_response_{model_idx}_auto_vs_human"
                consistency = rubric_auto_vs_human.get(vs_key, "无人工评分")

                if consistency not in ("无人工评分", "No human score"):
                    all_total += 1
                    if consistency in ("一致", "Consistent"):
                        all_matches += 1

    if all_total > 0:
        overall_consistency_rate = f"{all_matches / all_total * 100:.1f}%"

    return {
        "models": models_list,
        "task_scores": task_scores,
        "macro_averages": macro_averages,
        "micro_averages": micro_averages,
        "per_model_consistency_rate": model_consistency_stats,
        "overall_consistency_rate": overall_consistency_rate,
    }
